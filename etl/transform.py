from __future__ import annotations

import argparse
import csv
import datetime
import re
import sys
from collections import defaultdict
from pathlib import Path

if sys.path:
    script_dir = Path(__file__).resolve().parent
    if Path(sys.path[0]).resolve() == script_dir:
        sys.path[0] = str(script_dir.parent)

from dataclasses import dataclass

import yaml
from openpyxl import load_workbook

from etl.config import (
    DATASETS_BY_PROGRAM_ID,
    PARSE_PLANS_DIR,
    get_output_path,
    iter_parse_plan_files,
)
from etl.datasets.usda import HIERARCHY_COLUMNS as USDA_HIERARCHY_COLUMNS
from etl.datasets.usda import split_hierarchy as _usda_hierarchy

YEAR_RE = re.compile(r"(19|20)\d{2}")
NUMBER_RE = re.compile(r"^\(?[$]?\s*[-+]?\d[\d,]*(?:\.\d+)?\)?$")
# Strict pattern for fiscal-year column header cells.  The year must be the
# primary content: allows an optional "FY"/"fiscal year" prefix and a short
# trailing qualifier (e.g. "actual", "est.", "projected"), but rejects
# publication-date strings such as "February 2021" or "2021 to 2031".
YEAR_LABEL_RE = re.compile(
    r"^\s*(?:fy\s*|f\.y\.\s*|fiscal\s+year\s*)?((19|20)\d{2})"
    r"(?:\s*(?:actual|est(?:imate[ds]?)?\.?|proj(?:ect(?:ed|ion)?)?\.?|baseline))?"
    r"\s*$",
    re.IGNORECASE,
)
SHORT_YEAR_RANGE_RE = re.compile(
    r"^\s*((?:19|20)\d{2})\s*[-–—]\s*(\d{2})\s*$",
    re.IGNORECASE,
)
PERIOD_DECLARATION_RE = re.compile(
    r"\b(calendar|fiscal|award|school)\s+year\s+((?:19|20)\d{2})"
    r"(?:\s*[-–—]\s*((?:19|20)?\d{2}))?\b",
    re.IGNORECASE,
)
# Unit-declaration detection: section headers that announce the unit for the
# rows that follow.  CBO workbooks often contain multiple sub-sections per
# sheet (e.g. budget in "Billions of dollars" followed by enrollment in
# "Millions of people" followed by per-enrollee figures in "Dollars").  These
# patterns match either a standalone unit line or a unit inside a parenthetical
# at the end of a section header (with an optional trailing footnote letter).
_UNIT_PHRASE = (
    r"(?:billions?|millions?|thousands?|trillions?|hundreds?)\s+of\s+"
    r"(?:dollars?|people|beneficiaries|enrollees|recipients|households|"
    r"borrowers|loans|workers|cases|claims|jobs|hours|units|barrels|tons|"
    r"adults|children|students|veterans|survivors|participants|families|awards)"
    r"|dollars\s+per\s+\w+(?:\s+\w+)?"  # "Dollars per enrollee", "Dollars per recipient"
    r"|percent(?:age)?(?:\s+of\s+\w+(?:\s+\w+)?)?"
    r"|number\s+of\s+\w+(?:\s+\w+)?"
    r"|(?:billions?|millions?|thousands?|trillions?)"  # standalone quantity, e.g. "(Thousands)"
)
UNIT_LINE_RE = re.compile(
    rf"^\s*(?:\(\s*)?({_UNIT_PHRASE})\b[^()]*$",
    re.IGNORECASE,
)
# Comma-qualified suffix (e.g. ", calendar year 2018") is consumed but NOT
# captured so the group returns only the clean unit phrase.
UNIT_PAREN_RE = re.compile(
    rf"\(\s*({_UNIT_PHRASE})\s*(?:,\s*[^()]+?)?\s*\)\s*[a-z]?\s*$",
    re.IGNORECASE,
)
# Standalone "Dollars" or "(Dollars)" with optional footnote marker.
DOLLARS_LINE_RE = re.compile(r"^\s*\(?\s*dollars\s*\)?\s*[a-z]?\s*$", re.IGNORECASE)
DOLLARS_PAREN_RE = re.compile(r"\(\s*dollars\s*\)\s*[a-z]?\s*$", re.IGNORECASE)
# Per-capita qualifiers that differentiate "Dollars per enrollee" from plain
# "Dollars". Matches the noun that follows "per" in section headers like
# "Average Federal Spending per Enrollee (Dollars)" or
# "Cost per Beneficiary (Dollars)".
_PER_QUALIFIER_RE = re.compile(
    r"\bper\s+(enrollee|beneficiar(?:y|ies)|participant|recipient|person|worker|capita)\b",
    re.IGNORECASE,
)
# Maximum rows to scan beyond the declared header when searching for year labels.
MAX_YEAR_SCAN_ROWS = 30
# When parse-plan year_columns are absent, scan only a bounded number of leading
# columns. CBO fiscal-year tables are left-aligned and this avoids O(max_column)
# scans on very wide worksheets.
MAX_INFERRED_YEAR_COLUMNS = 30
HEALTH_KEYWORDS = ("health", "medicare", "medicaid", "chip")
INCOME_SECURITY_KEYWORDS = (
    "child_support",
    "childsupport",
    "csec",
    "foster_care",
    "fostercare",
    "military_retirement",
    "militaryretirement",
    "snap",
    "social_security",
    "socialsecurity",
    "ssi",
    "student_loan",
    "studentloan",
    "tanf",
    "unemployment",
)
SLICE_KEYWORDS = {
    "health": HEALTH_KEYWORDS,
    "income-security": INCOME_SECURITY_KEYWORDS,
}
SLICE_CHOICES = ("health", "income-security", "remaining-programs", "all")
# Source workbooks contain useful historical actuals well before the baseline
# publication year.  The bounds are deliberately broad and are used only for
# recognizing year headers, not for deciding which observations deserve to be
# retained.
PLAUSIBLE_YEAR_MIN = 1900
PLAUSIBLE_YEAR_MAX = 2100

OUTPUT_COLUMNS = [
    "program",
    "category",
    "fiscal_year",
    "value",
    "unit",
    "source_file",
    "source_sheet",
    "is_total",
    "program_id",
    "category_path",
    "period_type",
    "period_start_year",
    "period_end_year",
    "period_label",
    "source_row",
    "source_column",
]
_CATEGORY_PATH_INDEX = OUTPUT_COLUMNS.index("category_path") + 1
USDA_OUTPUT_COLUMNS = (
    OUTPUT_COLUMNS[:_CATEGORY_PATH_INDEX]
    + USDA_HIERARCHY_COLUMNS
    + OUTPUT_COLUMNS[_CATEGORY_PATH_INDEX:]
)


CANONICAL_PROGRAMS = {
    program_id: dataset.title
    for program_id, dataset in DATASETS_BY_PROGRAM_ID.items()
}


@dataclass(frozen=True)
class PeriodColumn:
    column: int
    period_type: str
    start_year: int | None
    end_year: int | None
    label: str


@dataclass(frozen=True)
class PeriodBlock:
    header_row: int
    periods: tuple[PeriodColumn, ...]


@dataclass(frozen=True)
class SheetPlan:
    workbook: str
    sheet: str
    include: bool
    output_dataset: str
    header_end_row: int
    first_data_row: int | None
    year_columns: list[int]
    unit: str
    verification_exempt: bool = False
    period_type: str = "fiscal_year"


def _header_end_row(header_rows: str) -> int:
    if "-" in header_rows:
        _, end = header_rows.split("-", 1)
        return int(end)
    return int(header_rows)


def _to_text(value: object) -> str:
    if value is None:
        return ""
    # Excel labels frequently contain embedded line breaks and alignment-only
    # whitespace. Collapse them so each CSV observation remains a physical row.
    return " ".join(str(value).split())


def _parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _to_text(value)
    if not text:
        return None
    if not NUMBER_RE.match(text):
        return None
    normalized = text.replace("$", "").replace(",", "").replace(" ", "")
    negative = normalized.startswith("(") and normalized.endswith(")")
    normalized = normalized.strip("()")
    parsed = float(normalized)
    return -parsed if negative else parsed


def _is_total(category: str) -> bool:
    lowered = category.lower()
    return "total" in lowered or "subtotal" in lowered


def _looks_like_note(category: str) -> bool:
    lowered = category.lower()
    return lowered.startswith("note") or lowered.startswith("source")


def _begins_note_block(text_cells: list[tuple[int, str]]) -> bool:
    """Return whether a row starts a trailing notes/source section.

    A standalone ``Note:`` or ``Source:`` label often precedes rows containing
    numeric examples. Those numbers are explanatory prose, not observations
    under the preceding period header.
    """

    return any(
        column <= 3 and re.match(r"^(?:notes?|sources?)\s*[:.]", text, re.IGNORECASE)
        for column, text in text_cells
    )


def _get_row_category(worksheet, row: int) -> str:
    """Return the category text for *row*.

    Checks columns 1, 2, and 3 in order, returning the first non-empty value.
    This handles workbooks that use an indented multi-column layout where the
    category label appears in column 2 or 3 with earlier columns left blank.
    """
    for col in (1, 2, 3):
        cat = _to_text(worksheet.cell(row=row, column=col).value)
        if cat:
            return cat
    return ""


def _extract_unit_declaration(text: str) -> str | None:
    """Return the normalized unit string declared by *text*, or None.

    Detects three forms of unit declaration:

    1. A line whose primary content is a unit phrase, e.g.
       ``"Billions of dollars, by fiscal year"`` -> ``"Billions of dollars"``.
    2. A unit phrase inside a trailing parenthetical, e.g.
       ``"Federal Benefit Payments by Eligibility Category (Billions of dollars)"``
       -> ``"Billions of dollars"``.
    3. The standalone word ``"Dollars"`` (optionally with a footnote letter
       suffix), e.g. ``"Average Federal Spending per Enrollee (Dollars)c"``
       -> ``"Dollars"``.

    The returned string preserves the matched phrase's capitalization but
    strips footnote markers and surrounding punctuation so values are
    comparable downstream.
    """
    if not text:
        return None
    stripped = text.strip()
    if not stripped:
        return None
    # Form 3: standalone "Dollars" (line or paren); test first because it's
    # narrower than the generic phrase matcher.
    if DOLLARS_LINE_RE.match(stripped):
        return "Dollars"
    paren_dollars = DOLLARS_PAREN_RE.search(stripped)
    if paren_dollars:
        per_m = _PER_QUALIFIER_RE.search(stripped)
        if per_m:
            return f"Dollars per {per_m.group(1).lower()}"
        return "Dollars"
    # Form 2: unit phrase inside trailing parenthetical (e.g. section header).
    paren_match = UNIT_PAREN_RE.search(stripped)
    if paren_match:
        return paren_match.group(1).strip().rstrip(",").strip()
    # Form 1: line whose primary content is a unit phrase.
    line_match = UNIT_LINE_RE.match(stripped)
    if line_match:
        return line_match.group(1).strip().rstrip(",").strip()
    return None


def _normalize_unit_string(s: str) -> str:
    """Normalize a raw unit label (e.g. from the parse-plan YAML) to a consistent
    unit string.

    Handles embedded phrases such as ``'By Fiscal Year, Billions of Dollars'``
    or ``'Outlays in Millions of Dollars, by Fiscal Year'`` by finding the unit
    phrase anywhere in the string.  Returns the phrase with first-letter
    capitalisation and the remainder lower-cased for consistency across vintages.
    """
    if not s:
        return s
    phrase = _extract_unit_declaration(s)
    if phrase is None:
        m = re.search(rf"({_UNIT_PHRASE})", s, re.IGNORECASE)
        phrase = m.group(1).strip() if m else None
    if phrase:
        # Plain-cell reads concatenate a superscript footnote marker with the
        # final noun (for example, ``Number of Beneficiaries`` + superscript
        # ``e`` becomes ``Number of Beneficiariese``). Remove that marker only
        # from the closed set of nouns supported by the unit parser.
        phrase = re.sub(
            r"\b(beneficiaries|enrollees|recipients|households|borrowers|loans|"
            r"workers|cases|claims|jobs|hours|units|barrels|tons|adults|children|"
            r"students|veterans|survivors|participants|families|awards)([a-z])$",
            r"\1",
            phrase,
            flags=re.IGNORECASE,
        )
        return phrase[:1].upper() + phrase[1:].lower()
    return s.strip()


def _program_id(filename: str) -> str:
    """Return the stable CBO source identifier at the start of a filename."""

    match = re.match(r"^(\d+)-", Path(filename).name)
    return match.group(1) if match else ""


def _canonical_program_name(filename: str) -> str:
    """Return a stable program label, falling back to the legacy inference."""

    return CANONICAL_PROGRAMS.get(_program_id(filename), _infer_program_name(filename))


def _exact_year(value: object) -> int | None:
    """Recognize a cell whose complete value is a four-digit year."""

    if isinstance(value, (datetime.datetime, datetime.date, bool)) or value is None:
        return None
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            year = int(value)
            if PLAUSIBLE_YEAR_MIN <= year <= PLAUSIBLE_YEAR_MAX:
                return year
        return None
    match = YEAR_LABEL_RE.match(_to_text(value))
    if not match:
        return None
    year = int(match.group(1))
    return year if PLAUSIBLE_YEAR_MIN <= year <= PLAUSIBLE_YEAR_MAX else None


def _short_year_range(value: object) -> tuple[int, int] | None:
    """Recognize academic-style labels such as ``2024-25``."""

    if not isinstance(value, str):
        return None
    match = SHORT_YEAR_RANGE_RE.match(_to_text(value))
    if not match:
        return None
    start = int(match.group(1))
    end = (start // 100) * 100 + int(match.group(2))
    if end <= start:
        end += 100
    return (start, end) if end == start + 1 else None


def _range_period_type(worksheet, header_row: int) -> str:
    text = " ".join(
        value
        for row in range(max(1, header_row - 8), header_row + 1)
        for _, value in _row_text_cells(worksheet, row)
    ).lower()
    return "school_year" if "school year" in text else "award_year"


def _period_declaration(text_cells: list[tuple[int, str]]) -> tuple[str, int, int, str] | None:
    """Extract an explicitly named period from prose or a section label."""

    for _, text in text_cells:
        match = PERIOD_DECLARATION_RE.search(text)
        if not match:
            continue
        period_type = f"{match.group(1).lower()}_year"
        start = int(match.group(2))
        end_text = match.group(3)
        if end_text:
            if len(end_text) == 2:
                end = (start // 100) * 100 + int(end_text)
                if end <= start:
                    end += 100
            else:
                end = int(end_text)
        else:
            end = start
        return period_type, start, end, str(start) if start == end else f"{start}-{end}"
    return None


def _longest_consecutive_run(year_cells: list[tuple[int, int]]) -> int:
    """Return the longest adjacent-column run whose years increase by one."""

    longest = current = 0
    previous: tuple[int, int] | None = None
    for item in sorted(year_cells):
        if previous and item[0] == previous[0] + 1 and item[1] == previous[1] + 1:
            current += 1
        else:
            current = 1
        longest = max(longest, current)
        previous = item
    return longest


def _candidate_header_rows(worksheet, plan: SheetPlan) -> list[int]:
    """Find coordinated year-header rows rather than scanning columns alone.

    A real header normally contains a consecutive run of years across adjacent
    columns.  Requiring that pattern prevents ordinary data values such as
    1,919 or 2,030 from being mistaken for column headers.
    """

    planned = set(plan.year_columns)
    columns = sorted(planned) if planned else list(range(1, min(worksheet.max_column, 60) + 1))
    initial_header_limit = max(
        plan.header_end_row,
        (plan.first_data_row - 1) if plan.first_data_row and plan.first_data_row > 1 else 0,
        min(MAX_YEAR_SCAN_ROWS, worksheet.max_row),
    )
    candidates: list[int] = []
    for row in range(1, worksheet.max_row + 1):
        year_cells = []
        for column in columns:
            value = worksheet.cell(row, column).value
            year = _exact_year(value)
            short_range = _short_year_range(value)
            if year is not None:
                year_cells.append((column, year))
            elif short_range is not None:
                year_cells.append((column, short_range[0]))
        if not year_cells:
            continue
        longest = _longest_consecutive_run(year_cells)
        # Two-year and one-year tables are accepted only in the initial header
        # area and only when the parse plan identifies the relevant columns.
        is_header = longest >= 3
        if not is_header and row <= initial_header_limit:
            is_header = longest >= 2 or (len(planned) == 1 and len(year_cells) == 1)
        if is_header:
            candidates.append(row)
    return candidates


def _cumulative_start(worksheet, header_row: int, column: int, end_year: int) -> int | None:
    """Read stacked headers such as ``2025-`` above ``2029``."""

    for row in range(header_row - 1, max(0, header_row - 4), -1):
        text = _to_text(worksheet.cell(row=row, column=column).value)
        match = re.search(r"((?:19|20)\d{2})\s*[-–—]\s*$", text)
        if match:
            start = int(match.group(1))
            if start < end_year:
                return start
    return None


def _period_blocks(worksheet, plan: SheetPlan) -> list[PeriodBlock]:
    """Return every coordinated period block on a worksheet."""

    blocks: list[PeriodBlock] = []
    planned = set(plan.year_columns)
    columns = sorted(planned) if planned else list(range(1, min(worksheet.max_column, 60) + 1))
    for header_row in _candidate_header_rows(worksheet, plan):
        periods_by_column: dict[int, PeriodColumn] = {}
        for column in columns:
            header_value = worksheet.cell(header_row, column).value
            short_range = _short_year_range(header_value)
            if short_range is not None:
                start_year, end_year = short_range
                periods_by_column[column] = PeriodColumn(
                    column=column,
                    period_type=_range_period_type(worksheet, header_row),
                    start_year=start_year,
                    end_year=end_year,
                    label=f"{start_year}-{end_year}",
                )
                continue

            end_year = _exact_year(header_value)
            if end_year is not None:
                start_year = _cumulative_start(worksheet, header_row, column, end_year)
                if start_year is not None:
                    periods_by_column[column] = PeriodColumn(
                        column=column,
                        period_type="cumulative_fiscal_years",
                        start_year=start_year,
                        end_year=end_year,
                        label=f"{start_year}-{end_year}",
                    )
                else:
                    periods_by_column[column] = PeriodColumn(
                        column=column,
                        period_type=plan.period_type,
                        start_year=end_year,
                        end_year=end_year,
                        label=str(end_year),
                    )
                continue

            # Some vintages put complete cumulative labels (for example,
            # ``2023-2027``) in a row immediately above the annual-year row.
            for candidate_row in range(header_row, max(0, header_row - 4), -1):
                text = _to_text(worksheet.cell(candidate_row, column).value)
                match = re.fullmatch(r"\s*((?:19|20)\d{2})\s*[-–—]\s*((?:19|20)\d{2})\s*", text)
                if match and int(match.group(1)) < int(match.group(2)):
                    periods_by_column[column] = PeriodColumn(
                        column=column,
                        period_type="cumulative_fiscal_years",
                        start_year=int(match.group(1)),
                        end_year=int(match.group(2)),
                        label=f"{match.group(1)}-{match.group(2)}",
                    )
                    break

        if not periods_by_column:
            continue

        # Split blocks only on true blank-column gaps. Cumulative columns are
        # adjacent to annual columns and intentionally remain in the same block.
        groups: list[list[PeriodColumn]] = []
        for period in [periods_by_column[column] for column in sorted(periods_by_column)]:
            if not groups or period.column > groups[-1][-1].column + 1:
                groups.append([period])
            else:
                groups[-1].append(period)
        for group in groups:
            if len(group) < 2 and len(columns) != 1:
                continue
            blocks.append(PeriodBlock(header_row=header_row, periods=tuple(group)))
    return blocks


def _period_type_for_row(default_type: str, text: str) -> str:
    """Apply row-level period semantics explicitly stated by the source."""

    if default_type == "cumulative_fiscal_years":
        return default_type
    lowered = text.lower()
    if "calendar year" in lowered:
        return "calendar_year"
    if "award year" in lowered:
        return "award_year"
    if "school year" in lowered:
        return "school_year"
    return default_type


def _is_meaningful_text(text: str) -> bool:
    if not text:
        return False
    stripped = text.strip()
    if not stripped or re.fullmatch(r"[_—–-]+", stripped):
        return False
    lowered = stripped.lower()
    if lowered in {"continued", "congressional budget office"}:
        return False
    if re.fullmatch(r"cbo.?s\s+.+\s+baseline", lowered):
        return False
    if re.fullmatch(r"(?:january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{4}", lowered):
        return False
    return True


def _row_text_cells(worksheet, row: int, max_column: int | None = None) -> list[tuple[int, str]]:
    limit = min(worksheet.max_column, max_column or worksheet.max_column, 80)
    cells: list[tuple[int, str]] = []
    for column in range(1, limit + 1):
        value = worksheet.cell(row=row, column=column).value
        if isinstance(value, str):
            text = _to_text(value)
            if _is_meaningful_text(text):
                cells.append((column, text))
    return cells


def _row_unit(text_cells: list[tuple[int, str]]) -> str | None:
    detected: list[str] = []
    for _, text in text_cells:
        unit = _extract_unit_declaration(text)
        if unit:
            detected.append(_normalize_unit_string(unit))
    return detected[-1] if detected else None


def _unit_for_category(category_path: str, current_unit: str) -> str:
    """Use an explicit metric label when it is more specific than sheet metadata."""

    explicit = _extract_unit_declaration(category_path)
    if explicit:
        return _normalize_unit_string(explicit)
    lowered = category_path.lower()
    if "%" in category_path or "interest rate" in lowered:
        return "Percent"
    return current_unit or "Unknown"


def _build_unit_map(worksheet, plan: SheetPlan, year_columns: list[int]) -> dict[int, str]:
    """Return a map of *row* -> unit string in effect for that row.

    Pre-scans the sheet row-by-row, updating the running unit whenever a row's
    column-A text contains a unit declaration (see :func:`_extract_unit_declaration`).
    Rows that contain numeric data in any year column are NOT treated as unit
    declarations even if their text matches a unit pattern, because legitimate
    data rows occasionally include unit-like words.

    The default starting unit is ``plan.unit`` (from the parse plan), so sheets
    with no mid-sheet unit changes preserve their existing behavior.
    """
    row_unit: dict[int, str] = {}
    current = plan.unit
    for row in range(1, worksheet.max_row + 1):
        text_a = _to_text(worksheet.cell(row=row, column=1).value)
        if text_a:
            has_numeric_data = any(
                _parse_number(worksheet.cell(row=row, column=col).value) is not None
                for col in year_columns
            )
            if not has_numeric_data:
                detected = _extract_unit_declaration(text_a)
                if detected:
                    current = detected
        row_unit[row] = current
    return row_unit


def _extract_years(worksheet, plan: SheetPlan) -> dict[int, int]:
    """Return the annual fiscal-year columns in the first detected block.

    The transformation itself consumes :func:`_period_blocks` so cumulative
    and non-fiscal periods are preserved without overloading ``fiscal_year``.
    """

    for block in _period_blocks(worksheet, plan):
        years = {
            period.column: period.end_year
            for period in block.periods
            if period.period_type == "fiscal_year" and period.end_year is not None
        }
        if years:
            return years
    return {}


def _infer_first_data_row(worksheet, plan: SheetPlan, year_columns: list[int] | None = None) -> int | None:
    active_year_columns = year_columns if year_columns is not None else plan.year_columns
    if not active_year_columns:
        return None
    for row in range(plan.header_end_row + 1, worksheet.max_row + 1):
        category = _get_row_category(worksheet, row)
        if not category:
            continue
        if _looks_like_note(category):
            continue
        parsed_values = [_parse_number(worksheet.cell(row=row, column=col).value) for col in active_year_columns]
        has_value = any(value is not None for value in parsed_values)
        if has_value:
            return row
    return None


def _infer_program_name(filename: str) -> str:
    """Infer a display program name from CBO workbook filenames.

    Typical filenames follow `<id>-<year>-<month>-<program>.xlsx`, e.g.
    `51293-2024-06-childnutrition.xlsx`. If this shape is not present, the
    entire stem is used.
    """
    stem = Path(filename).stem.replace("_", "-")
    parts = stem.split("-")
    if len(parts) >= 4:
        name = "-".join(parts[3:])
    else:
        name = stem
    return " ".join(token for token in re.split(r"[-\s]+", name) if token).title()


def _find_sheet(sheetnames: list[str], target: str) -> str | None:
    """Return the actual sheet name from *sheetnames* that matches *target*.

    Tries an exact match first, then falls back to a whitespace-stripped
    comparison so that sheet names with incidental trailing spaces (a common
    Excel authoring artefact) are handled transparently.
    """
    if target in sheetnames:
        return target
    target_stripped = target.strip()
    for name in sheetnames:
        if name.strip() == target_stripped:
            return name
    return None


def _read_plan(path: Path) -> list[SheetPlan]:
    plans: list[SheetPlan] = []
    for plan_file in iter_parse_plan_files(path):
        payload = yaml.safe_load(plan_file.read_text(encoding="utf-8")) or {}
        for workbook_entry in payload.get("workbooks", []):
            workbook_name = workbook_entry.get("workbook")
            for sheet_entry in workbook_entry.get("sheets", []):
                plans.append(
                    SheetPlan(
                        workbook=workbook_name,
                        sheet=sheet_entry.get("sheet"),
                        include=bool(sheet_entry.get("include")),
                        output_dataset=sheet_entry.get("output_dataset", ""),
                        header_end_row=_header_end_row(str(sheet_entry.get("header_rows", "1-1"))),
                        first_data_row=sheet_entry.get("first_data_row"),
                        year_columns=[int(column) for column in sheet_entry.get("year_columns", [])],
                        unit=_normalize_unit_string(str(sheet_entry.get("unit", "")).strip()),
                        verification_exempt=bool(sheet_entry.get("verification_exempt", False)),
                        period_type=str(sheet_entry.get("period_type", "fiscal_year")).strip()
                        or "fiscal_year",
                    )
                )
    return plans


def _in_slice(plan: SheetPlan, slice_name: str) -> bool:
    if slice_name == "all":
        return True
    dataset = plan.output_dataset.lower()
    if slice_name == "remaining-programs":
        return not any(kw in dataset for kw in HEALTH_KEYWORDS) and not any(
            kw in dataset for kw in INCOME_SECURITY_KEYWORDS
        )
    keywords = SLICE_KEYWORDS.get(slice_name, ())
    return any(keyword in dataset for keyword in keywords)


def _write_dataset(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = USDA_OUTPUT_COLUMNS if rows and "table_title" in rows[0] else OUTPUT_COLUMNS
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def _append_error(errors: list[str], workbook: str, sheet: str, reason: str) -> None:
    errors.append(f"{workbook}\t{sheet}\t{reason}")


def _category_path(section_by_column: dict[int, str], category: str, category_column: int) -> str:
    """Combine structural headings with a leaf label without duplicating text."""

    parts: list[str] = []
    for column, heading in sorted(section_by_column.items()):
        if column > category_column:
            continue
        normalized = heading.strip()
        if normalized and (not parts or normalized.casefold() != parts[-1].casefold()):
            parts.append(normalized)
    if not parts or category.casefold() != parts[-1].casefold():
        parts.append(category)
    return " / ".join(parts)


def _combined_category_path(
    structural: dict[int, str], inline: dict[int, str], category: str, category_column: int
) -> str:
    """Build a path when structural and inline labels share indentation columns."""

    parts: list[str] = []
    for headings, include_same_column in ((structural, True), (inline, False)):
        for column, heading in sorted(headings.items()):
            if column > category_column or (column == category_column and not include_same_column):
                continue
            normalized = heading.strip()
            if normalized and normalized.casefold() not in {part.casefold() for part in parts}:
                parts.append(normalized)
    if category.casefold() not in {part.casefold() for part in parts}:
        parts.append(category)
    return " / ".join(parts)


def _record(
    plan: SheetPlan,
    category: str,
    category_path: str,
    period_type: str,
    start_year: int | None,
    end_year: int | None,
    period_label: str,
    value: float,
    unit: str,
    row: int,
    column: int,
) -> dict:
    """Build one normalized observation with complete source-cell lineage."""

    program_id = _program_id(plan.workbook)
    fiscal_year: int | str = end_year if period_type == "fiscal_year" and end_year is not None else ""
    record = {
        "program": _canonical_program_name(plan.workbook),
        "category": category,
        "fiscal_year": fiscal_year,
        "value": value,
        "unit": unit or "Unknown",
        "source_file": plan.workbook,
        "source_sheet": plan.sheet,
        "is_total": str(_is_total(category_path)).lower(),
        "program_id": program_id,
        "category_path": category_path,
    }
    if program_id == "51317":
        record.update(_usda_hierarchy(category_path, category))
    record.update({
        "period_type": period_type,
        "period_start_year": start_year if start_year is not None else "",
        "period_end_year": end_year if end_year is not None else "",
        "period_label": period_label,
        "source_row": row,
        "source_column": column,
    })
    return record


def _heading_candidates(text_cells: list[tuple[int, str]], period_start_column: int) -> list[tuple[int, str]]:
    candidates: list[tuple[int, str]] = []
    for column, text in text_cells:
        if column >= period_start_column or _looks_like_note(text):
            continue
        if _parse_number(text) is not None or _exact_year(text) is not None:
            continue
        if _extract_unit_declaration(text) and re.fullmatch(r"\s*\(?\s*" + _UNIT_PHRASE + r"[^)]*\)?\s*", text, re.IGNORECASE):
            continue
        candidates.append((column, text))
    return candidates


def _update_sections(section_by_column: dict[int, str], headings: list[tuple[int, str]]) -> None:
    for column, heading in headings:
        for existing_column in [key for key in section_by_column if key >= column]:
            del section_by_column[existing_column]
        section_by_column[column] = heading


def _nearest_table_heading(worksheet, header_row: int, period_start_column: int) -> str:
    """Return the closest descriptive title immediately above a period header."""

    for row in range(header_row - 1, max(0, header_row - 9), -1):
        candidates = _heading_candidates(_row_text_cells(worksheet, row), period_start_column)
        for _, text in reversed(candidates):
            lowered = text.lower()
            if lowered in {"actual", "estimated", "projected", "actual and projected"}:
                continue
            if "baseline" in lowered and ("cbo" in lowered or "budget office" in lowered):
                continue
            return text
    return ""


def _records_from_period_table(worksheet, plan: SheetPlan, blocks: list[PeriodBlock]) -> list[dict]:
    """Parse column-oriented tables while preserving headings and period types."""

    blocks_by_row: dict[int, list[PeriodColumn]] = defaultdict(list)
    for block in blocks:
        blocks_by_row[block.header_row].extend(block.periods)

    current_periods: list[PeriodColumn] = []
    current_unit = plan.unit
    section_by_column: dict[int, str] = {}
    inline_by_column: dict[int, str] = {}
    pending_label: tuple[int, str, int] | None = None
    records: list[dict] = []
    minimum_data_row = plan.first_data_row or min(block.header_row for block in blocks) + 1
    first_period_column = min(period.column for block in blocks for period in block.periods)
    table_heading = ""
    in_note_block = False

    for row in range(1, worksheet.max_row + 1):
        text_cells = _row_text_cells(worksheet, row)
        detected_unit = _row_unit(text_cells)
        if detected_unit:
            current_unit = detected_unit

        if row in blocks_by_row:
            unique = {period.column: period for period in blocks_by_row[row]}
            current_periods = [unique[column] for column in sorted(unique)]
            table_heading = _nearest_table_heading(worksheet, row, min(unique)) or table_heading
            section_by_column.clear()
            inline_by_column.clear()
            pending_label = None
            in_note_block = False
            continue
        if row >= minimum_data_row and _begins_note_block(text_cells):
            in_note_block = True
            pending_label = None
            continue
        if in_note_block:
            continue
        if not current_periods:
            if row >= minimum_data_row:
                headings = _heading_candidates(text_cells, first_period_column)
                if headings:
                    _update_sections(section_by_column, headings)
                    inline_by_column.clear()
            continue
        if row < minimum_data_row:
            period_start_column = min(period.column for period in current_periods)
            headings = _heading_candidates(text_cells, period_start_column)
            if headings:
                _update_sections(section_by_column, headings)
                inline_by_column.clear()
            continue

        period_start_column = min(period.column for period in current_periods)
        numeric_periods = [
            (period, value)
            for period in current_periods
            if (value := _parse_number(worksheet.cell(row=row, column=period.column).value)) is not None
        ]
        headings = _heading_candidates(text_cells, period_start_column)

        if not numeric_periods:
            if headings:
                _update_sections(section_by_column, headings)
                inline_by_column.clear()
                pending_label = (*headings[-1], row)
            continue

        label_candidates = headings
        if label_candidates:
            category_column, category = label_candidates[-1]
        elif pending_label and row - pending_label[2] <= 2:
            category_column, category, _ = pending_label
        elif table_heading:
            category_column, category = 1, table_heading
        else:
            category_column, category = 1, f"Unlabeled value (row {row})"
        if _looks_like_note(category):
            continue

        if len(label_candidates) > 1:
            _update_sections(inline_by_column, label_candidates[:-1])
        category_path = _combined_category_path(section_by_column, inline_by_column, category, category_column)
        if table_heading and not category_path.casefold().startswith(table_heading.casefold()):
            category_path = f"{table_heading} / {category_path}"
        explicit_unit = _row_unit(label_candidates)
        unit = explicit_unit or _unit_for_category(category_path, current_unit)
        for period, value in numeric_periods:
            period_type = _period_type_for_row(period.period_type, category_path)
            records.append(
                _record(
                    plan=plan,
                    category=category,
                    category_path=category_path,
                    period_type=period_type,
                    start_year=period.start_year,
                    end_year=period.end_year,
                    period_label=period.label,
                    value=value,
                    unit=unit,
                    row=row,
                    column=period.column,
                )
            )
        pending_label = None
    return records


def _column_heading(worksheet, row: int, column: int) -> str:
    """Find the nearest text header above a value in a nonstandard table."""

    for candidate_row in range(row - 1, max(0, row - 12), -1):
        text = _to_text(worksheet.cell(candidate_row, column).value)
        if text and _is_meaningful_text(text) and _parse_number(text) is None:
            return text
    return ""


def _generic_column_contexts(worksheet, first_data_row: int, max_column: int) -> dict[int, str]:
    """Build scenario/estimate labels for columns in comparison-style tables."""

    contexts: dict[int, str] = {}
    for column in range(1, max_column + 1):
        parts: list[str] = []
        for row in range(1, first_data_row):
            text = _to_text(worksheet.cell(row, column).value)
            lowered = text.lower()
            if not text or not _is_meaningful_text(text) or len(text) > 80:
                continue
            if _extract_unit_declaration(text) or _period_declaration([(column, text)]):
                continue
            if any(
                phrase in lowered
                for phrase in (
                    "congressional budget office",
                    "baseline projections",
                    "selected categories",
                    "table ",
                )
            ):
                continue
            if text.casefold() not in {part.casefold() for part in parts}:
                parts.append(text.replace("\n", " "))
        if parts:
            contexts[column] = " ".join(parts[-2:])
    return contexts


def _records_from_generic_table(worksheet, plan: SheetPlan) -> list[dict]:
    """Represent nonstandard included sheets without inventing annual semantics.

    Every emitted measure retains its exact source coordinate.  A leading year
    key is used when present; otherwise the period is explicitly ``unmapped``.
    """

    first_row = plan.first_data_row or plan.header_end_row + 1
    max_column = min(worksheet.max_column, 200)
    current_unit = plan.unit
    current_period: tuple[str, int, int, str] | None = None
    section_by_column: dict[int, str] = {}
    inline_by_column: dict[int, str] = {}
    records: list[dict] = []
    column_contexts = _generic_column_contexts(worksheet, first_row, max_column)
    in_note_block = False

    for row in range(1, worksheet.max_row + 1):
        text_cells = _row_text_cells(worksheet, row, max_column)
        if row >= first_row and _begins_note_block(text_cells):
            in_note_block = True
            continue
        if in_note_block:
            continue
        detected_unit = _row_unit(text_cells)
        if detected_unit:
            current_unit = detected_unit
        declared_period = _period_declaration(text_cells)
        if declared_period:
            current_period = declared_period
        if row < first_row:
            continue

        numeric_cells = [
            (column, value)
            for column in range(1, max_column + 1)
            if (value := _parse_number(worksheet.cell(row=row, column=column).value)) is not None
        ]
        if not numeric_cells:
            headings = _heading_candidates(text_cells, max_column + 1)
            if headings:
                _update_sections(section_by_column, headings)
                inline_by_column.clear()
            continue

        leading_year: tuple[int, int] | None = None
        for column, _ in numeric_cells:
            year = _exact_year(worksheet.cell(row=row, column=column).value)
            if year is not None and column <= 3:
                leading_year = (column, year)
                break

        for column, value in numeric_cells:
            if leading_year and column == leading_year[0]:
                continue
            labels = [item for item in text_cells if item[0] < column and _parse_number(item[1]) is None]
            if labels:
                category_column, category = labels[-1]
            else:
                header = _column_heading(worksheet, row, column)
                category_column = column
                category = header or f"Unlabeled value (row {row}, column {column})"
            if _looks_like_note(category):
                continue
            if labels and len(labels) > 1:
                _update_sections(inline_by_column, labels[:-1])
            category_path = _combined_category_path(section_by_column, inline_by_column, category, category_column)
            if leading_year:
                period_type = _period_type_for_row("fiscal_year", category_path)
                start_year = end_year = leading_year[1]
                period_label = str(leading_year[1])
            elif current_period:
                period_type, start_year, end_year, period_label = current_period
            else:
                period_type = "unmapped"
                start_year = end_year = None
                period_label = "Not identified in source headers"
            column_context = column_contexts.get(column, "")
            if column_context and column_context.casefold() not in category_path.casefold():
                category_path = f"{category_path} / {column_context}"
            records.append(
                _record(
                    plan=plan,
                    category=category,
                    category_path=category_path,
                    period_type=period_type,
                    start_year=start_year,
                    end_year=end_year,
                    period_label=period_label,
                    value=value,
                    unit=_unit_for_category(category_path, current_unit),
                    row=row,
                    column=column,
                )
            )
    return records


def _records_for_sheet(worksheet, plan: SheetPlan) -> tuple[list[dict], str | None]:
    blocks = _period_blocks(worksheet, plan)
    if blocks:
        return _records_from_period_table(worksheet, plan, blocks), None
    records = _records_from_generic_table(worksheet, plan)
    if records and all(record["period_type"] != "unmapped" for record in records):
        warning = "nonstandard comparison layout; used coordinate-preserving comparison parser"
    else:
        warning = "no coordinated period header; used coordinate-preserving generic parser"
    return records, warning


def run_transform(
    parse_plan_path: Path = PARSE_PLANS_DIR,
    input_dir: Path = Path("data/raw"),
    output_dir: Path = Path("data/processed"),
    slice_name: str = "health",
) -> int:
    if slice_name not in SLICE_CHOICES:
        raise ValueError(f"Unsupported slice: {slice_name}")
    plans = [plan for plan in _read_plan(parse_plan_path) if plan.include and _in_slice(plan, slice_name)]
    records_by_dataset: dict[str, list[dict]] = defaultdict(list)
    workbook_by_dataset: dict[str, str] = {}
    errors: list[str] = []
    warnings: list[str] = []

    plans_by_workbook: dict[str, list[SheetPlan]] = defaultdict(list)
    for plan in plans:
        plans_by_workbook[plan.workbook].append(plan)
        workbook_by_dataset.setdefault(plan.output_dataset, plan.workbook)

    for workbook_name, workbook_plans in plans_by_workbook.items():
        workbook_path = input_dir / workbook_name
        if not workbook_path.exists():
            for plan in workbook_plans:
                _append_error(errors, plan.workbook, plan.sheet, "workbook not found")
            continue
        workbook = load_workbook(workbook_path, data_only=True)
        try:
            for plan in workbook_plans:
                actual_sheet = _find_sheet(workbook.sheetnames, plan.sheet)
                if actual_sheet is None:
                    _append_error(errors, plan.workbook, plan.sheet, "sheet not found")
                    continue
                worksheet = workbook[actual_sheet]
                sheet_records, warning = _records_for_sheet(worksheet, plan)
                records_by_dataset[plan.output_dataset].extend(sheet_records)
                if warning:
                    warnings.append(f"{plan.workbook}\t{plan.sheet}\t{warning}")
                if not sheet_records:
                    _append_error(errors, plan.workbook, plan.sheet, "no data rows parsed")
        finally:
            workbook.close()

    output_dir.mkdir(parents=True, exist_ok=True)
    if slice_name == "all":
        for existing_csv in output_dir.rglob("*.csv"):
            existing_csv.unlink()
    for dataset, rows in records_by_dataset.items():
        output_path = get_output_path(
            output_dir,
            workbook=workbook_by_dataset[dataset],
            output_dataset=dataset,
        )
        _write_dataset(output_path, rows)

    error_path = output_dir / "parse_errors.log"
    error_body = "\n".join(errors)
    if errors:
        error_body += "\n"
    error_path.write_text(error_body, encoding="utf-8")
    warning_path = output_dir / "parse_warnings.log"
    warning_body = "\n".join(warnings)
    if warnings:
        warning_body += "\n"
    warning_path.write_text(warning_body, encoding="utf-8")
    print(
        f"Transform complete. slice={slice_name}, datasets={len(records_by_dataset)}, "
        f"rows={sum(len(rows) for rows in records_by_dataset.values())}, "
        f"errors={len(errors)}, warnings={len(warnings)}"
    )
    return 1 if errors else 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Transform CBO baseline workbooks into tidy CSV datasets.")
    parser.add_argument("--parse-plan", type=Path, default=PARSE_PLANS_DIR)
    parser.add_argument("--input-dir", type=Path, default=Path("data/raw"))
    parser.add_argument("--output-dir", type=Path, default=Path("data/processed"))
    parser.add_argument("--slice", dest="slice_name", choices=SLICE_CHOICES, default="health")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    return run_transform(
        parse_plan_path=args.parse_plan,
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        slice_name=args.slice_name,
    )


if __name__ == "__main__":
    raise SystemExit(main())
