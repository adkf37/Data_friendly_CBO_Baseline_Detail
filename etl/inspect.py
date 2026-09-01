from __future__ import annotations

import argparse
import json
import re
import zipfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

FISCAL_YEAR_RE = re.compile(r"^(?:fy\s*)?(19|20)\d{2}$", re.IGNORECASE)
UNIT_HINTS = ("million", "billion", "dollar", "percent", "share", "index")
NOTES_HINTS = ("note", "source", "totals may not sum", "n/a", "na")
METADATA_HINTS = ("contents", "about", "overview", "readme", "instructions")


def _normalize_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _detect_fiscal_year_columns(rows: Iterable[list[str]]) -> list[int]:
    hits: dict[int, int] = {}
    for row in rows:
        for index, cell in enumerate(row, start=1):
            if FISCAL_YEAR_RE.match(cell):
                hits[index] = hits.get(index, 0) + 1
    return sorted(column for column, count in hits.items() if count >= 1)


def _detect_unit_text(rows: Iterable[list[str]]) -> str | None:
    for row in rows:
        for cell in row:
            lowered = cell.lower()
            if any(hint in lowered for hint in UNIT_HINTS):
                return cell
    return None


def _row_numeric_count(row: list[str]) -> int:
    count = 0
    for cell in row:
        try:
            if cell:
                float(cell.replace(",", ""))
                count += 1
        except ValueError:
            pass
    return count


def _infer_first_data_row(rows: list[list[str]], fiscal_year_columns: list[int]) -> int | None:
    def cell_at(row: list[str], column: int) -> str:
        return row[column - 1] if column <= len(row) else ""

    for row_index, row in enumerate(rows, start=1):
        numeric_count = _row_numeric_count(row)
        has_label = bool(row and row[0])
        has_year_data = any(cell_at(row, column) for column in fiscal_year_columns)
        looks_like_header = bool(fiscal_year_columns) and all(
            FISCAL_YEAR_RE.match(cell_at(row, column) or "") for column in fiscal_year_columns
        )
        if looks_like_header:
            continue
        if has_label and numeric_count >= 1 and (has_year_data or numeric_count >= 2):
            return row_index
    return None


def _infer_header_rows(first_data_row: int | None) -> str:
    if first_data_row is None:
        return "1-1"
    if first_data_row <= 2:
        return "1-1"
    return f"1-{first_data_row - 1}"


def _classify_sheet(
    sheet_name: str,
    rows: list[list[str]],
    fiscal_year_columns: list[int],
    first_data_row: int | None,
) -> str:
    lowered_name = sheet_name.lower()
    all_text = " ".join(cell.lower() for row in rows for cell in row if cell)
    numeric_rows = sum(1 for row in rows if _row_numeric_count(row) >= 2)

    if any(hint in lowered_name for hint in ("note", "notes")) or any(hint in all_text for hint in NOTES_HINTS):
        if numeric_rows <= 1:
            return "notes"
    if any(hint in lowered_name for hint in METADATA_HINTS):
        return "metadata"
    if fiscal_year_columns or (first_data_row is not None and numeric_rows >= 1):
        return "data"
    return "unknown"


def _detect_multiple_tables(rows: list[list[str]], first_data_row: int | None) -> bool:
    if first_data_row is None:
        return False
    data_like_rows = [
        row_index
        for row_index, row in enumerate(rows, start=1)
        if row_index >= first_data_row and _row_numeric_count(row) >= 1 and any(row)
    ]
    if len(data_like_rows) < 3:
        return False
    gaps = [b - a for a, b in zip(data_like_rows, data_like_rows[1:])]
    return any(gap >= 3 for gap in gaps)


def _sheet_has_merged_cells(workbook_path: Path, worksheet_xml_path: str | None) -> bool:
    if not worksheet_xml_path:
        return False
    with zipfile.ZipFile(workbook_path) as archive:
        with archive.open(worksheet_xml_path) as sheet_xml:
            for chunk in iter(lambda: sheet_xml.read(4096), b""):
                if b"<mergeCell" in chunk:
                    return True
    return False


def profile_sheet(workbook_path: Path, sheet_name: str) -> dict:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        has_merged_cells = _sheet_has_merged_cells(workbook_path, getattr(worksheet, "_worksheet_path", None))
        row_count = worksheet.max_row or 0
        column_count = worksheet.max_column or 0

        rows: list[list[str]] = []
        for values in worksheet.iter_rows(min_row=1, max_row=row_count, max_col=column_count, values_only=True):
            normalized = [_normalize_cell(value) for value in values]
            rows.append(normalized)
    finally:
        workbook.close()

    fiscal_year_columns = _detect_fiscal_year_columns(rows[:10] if rows else [])
    unit_text = _detect_unit_text(rows[:20] if rows else [])
    likely_first_data_row = _infer_first_data_row(rows, fiscal_year_columns)
    likely_header_rows = _infer_header_rows(likely_first_data_row)
    classification = _classify_sheet(sheet_name, rows, fiscal_year_columns, likely_first_data_row)
    multiple_tables_flagged = _detect_multiple_tables(rows, likely_first_data_row)

    return {
        "sheet_name": sheet_name,
        "row_count": row_count,
        "column_count": column_count,
        "has_merged_cells": has_merged_cells,
        "likely_header_rows": likely_header_rows,
        "likely_first_data_row": likely_first_data_row,
        "fiscal_year_columns": fiscal_year_columns,
        "unit_text": unit_text,
        "classification": classification,
        "multiple_tables_flagged": multiple_tables_flagged,
    }


def inspect_workbook(workbook_path: Path) -> dict:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet_names = list(workbook.sheetnames)
    finally:
        workbook.close()

    profiles = [profile_sheet(workbook_path, sheet_name) for sheet_name in sheet_names]
    return {"workbook": workbook_path.name, "sheet_count": len(sheet_names), "sheets": profiles}


def render_report(profiles: list[dict]) -> str:
    lines: list[str] = []
    lines.append("# Workbook Inspection Report")
    lines.append("")
    lines.append("Generated by workbook inspection tooling.")
    lines.append("")
    if not profiles:
        lines.append("No `.xlsx` files were found in `data/raw/`.")
        lines.append("")
        return "\n".join(lines) + "\n"

    for workbook_profile in profiles:
        lines.append(f"## {workbook_profile['workbook']}")
        lines.append("")
        lines.append(f"- Sheet count: {workbook_profile['sheet_count']}")
        lines.append("")
        for sheet in workbook_profile["sheets"]:
            lines.append(f"### {sheet['sheet_name']}")
            lines.append("")
            lines.append(f"- Dimensions: {sheet['row_count']} rows x {sheet['column_count']} columns")
            lines.append(f"- Merged cells present: {'yes' if sheet['has_merged_cells'] else 'no'}")
            lines.append(f"- Likely header rows: {sheet['likely_header_rows']}")
            lines.append(
                f"- Likely first data row: {sheet['likely_first_data_row'] if sheet['likely_first_data_row'] is not None else 'unknown'}"
            )
            lines.append(
                f"- Fiscal-year columns: {', '.join(str(column) for column in sheet['fiscal_year_columns']) if sheet['fiscal_year_columns'] else 'none detected'}"
            )
            lines.append(f"- Detected unit text: {sheet['unit_text'] or 'none detected'}")
            lines.append(f"- Classification: `{sheet['classification']}`")
            lines.append(
                f"- Multiple logical tables flagged: {'yes' if sheet['multiple_tables_flagged'] else 'no'}"
            )
            lines.append("")
    lines.append("## Open Questions")
    lines.append("")
    lines.append("- Review any sheets flagged with multiple logical tables before transform planning.")
    lines.append("- Confirm unit strings where detection returned `none detected`.")
    lines.append("")
    lines.append("## Machine-Readable Summary")
    lines.append("")
    lines.append("```json")
    lines.append(json.dumps(profiles, indent=2))
    lines.append("```")
    lines.append("")
    return "\n".join(lines)


def run_inspection(input_dir: Path = Path("data/raw"), report_path: Path = Path("docs/inspection_report.md")) -> int:
    workbook_paths = []
    if input_dir.exists():
        workbook_paths = sorted(path for path in input_dir.glob("*.xlsx") if path.is_file())
    profiles = [inspect_workbook(workbook_path) for workbook_path in workbook_paths]

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(render_report(profiles), encoding="utf-8")
    print(f"Inspected {len(workbook_paths)} workbook(s). report={report_path}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect downloaded CBO Excel workbooks and produce an inspection report.")
    parser.add_argument("--input-dir", type=Path, default=Path("data/raw"), help="Directory containing downloaded .xlsx files")
    parser.add_argument("--output", type=Path, default=Path("docs/inspection_report.md"), help="Inspection report markdown output path")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    return run_inspection(input_dir=args.input_dir, report_path=args.output)


if __name__ == "__main__":
    raise SystemExit(main())
