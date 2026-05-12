from __future__ import annotations

import argparse
import csv
import datetime
import re
import sys
from collections import defaultdict
from pathlib import Path

if sys.path:
    script_dir = Path(__file__).resolve().parent
    if Path(sys.path[0]).resolve() == script_dir:
        sys.path[0] = str(script_dir.parent)

from dataclasses import dataclass

import yaml
from openpyxl import load_workbook

YEAR_RE = re.compile(r"(19|20)\d{2}")
NUMBER_RE = re.compile(r"^\(?[$]?\s*[-+]?\d[\d,]*(?:\.\d+)?\)?$")
# Strict pattern for fiscal-year column header cells.  The year must be the
# primary content: allows an optional "FY"/"fiscal year" prefix and a short
# trailing qualifier (e.g. "actual", "est.", "projected"), but rejects
# publication-date strings such as "February 2021" or "2021 to 2031".
YEAR_LABEL_RE = re.compile(
    r"^\s*(?:fy\s*|f\.y\.\s*|fiscal\s+year\s*)?((19|20)\d{2})"
    r"(?:\s*(?:actual|est(?:imate[ds]?)?\.?|proj(?:ect(?:ed|ion)?)?\.?|baseline))?"
    r"\s*$",
    re.IGNORECASE,
)
# Unit-declaration detection: section headers that announce the unit for the
# rows that follow.  CBO workbooks often contain multiple sub-sections per
# sheet (e.g. budget in "Billions of dollars" followed by enrollment in
# "Millions of people" followed by per-enrollee figures in "Dollars").  These
# patterns match either a standalone unit line or a unit inside a parenthetical
# at the end of a section header (with an optional trailing footnote letter).
_UNIT_PHRASE = (
    r"(?:billions?|millions?|thousands?|trillions?|hundreds?)\s+of\s+"
    r"(?:dollars?|people|beneficiaries|enrollees|recipients|households|"
    r"borrowers|loans|workers|cases|claims|jobs|hours|units|barrels|tons)"
    r"|dollars\s+per\s+\w+(?:\s+\w+)?"  # "Dollars per enrollee", "Dollars per recipient"
    r"|percent(?:age)?(?:\s+of\s+\w+(?:\s+\w+)?)?"
    r"|number\s+of\s+\w+(?:\s+\w+)?"
    r"|(?:billions?|millions?|thousands?|trillions?)"  # standalone quantity, e.g. "(Thousands)"
)
UNIT_LINE_RE = re.compile(
    rf"^\s*(?:\(\s*)?({_UNIT_PHRASE})\b[^()]*$",
    re.IGNORECASE,
)
# Comma-qualified suffix (e.g. ", calendar year 2018") is consumed but NOT
# captured so the group returns only the clean unit phrase.
UNIT_PAREN_RE = re.compile(
    rf"\(\s*({_UNIT_PHRASE})\s*(?:,\s*[^()]+?)?\s*\)\s*[a-z]?\s*$",
    re.IGNORECASE,
)
# Standalone "Dollars" or "(Dollars)" with optional footnote marker.
DOLLARS_LINE_RE = re.compile(r"^\s*\(?\s*dollars\s*\)?\s*[a-z]?\s*$", re.IGNORECASE)
DOLLARS_PAREN_RE = re.compile(r"\(\s*dollars\s*\)\s*[a-z]?\s*$", re.IGNORECASE)
# Maximum rows to scan beyond the declared header when searching for year labels.
MAX_YEAR_SCAN_ROWS = 30
# When parse-plan year_columns are absent, scan only a bounded number of leading
# columns. CBO fiscal-year tables are left-aligned and this avoids O(max_column)
# scans on very wide worksheets.
MAX_INFERRED_YEAR_COLUMNS = 30
HEALTH_KEYWORDS = ("health", "medicare", "medicaid", "chip")
INCOME_SECURITY_KEYWORDS = (
    "child_support",
    "childsupport",
    "csec",
    "foster_care",
    "fostercare",
    "military_retirement",
    "militaryretirement",
    "snap",
    "social_security",
    "socialsecurity",
    "ssi",
    "student_loan",
    "studentloan",
    "tanf",
    "unemployment",
)
SLICE_KEYWORDS = {
    "health": HEALTH_KEYWORDS,
    "income-security": INCOME_SECURITY_KEYWORDS,
}
SLICE_CHOICES = ("health", "income-security", "remaining-programs", "all")
PLAUSIBLE_YEAR_MIN = 2019
PLAUSIBLE_YEAR_MAX = 2040

OUTPUT_COLUMNS = [
    "program",
    "category",
    "fiscal_year",
    "value",
    "unit",
    "source_file",
    "source_sheet",
    "is_total",
]


@dataclass(frozen=True)
class SheetPlan:
    workbook: str
    sheet: str
    include: bool
    output_dataset: str
    header_end_row: int
    first_data_row: int | None
    year_columns: list[int]
    unit: str
    verification_exempt: bool = False


def _header_end_row(header_rows: str) -> int:
    if "-" in header_rows:
        _, end = header_rows.split("-", 1)
        return int(end)
    return int(header_rows)


def _to_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _to_text(value)
    if not text:
        return None
    if not NUMBER_RE.match(text):
        return None
    normalized = text.replace("$", "").replace(",", "").replace(" ", "")
    negative = normalized.startswith("(") and normalized.endswith(")")
    normalized = normalized.strip("()")
    parsed = float(normalized)
    return -parsed if negative else parsed


def _is_total(category: str) -> bool:
    lowered = category.lower()
    return "total" in lowered or "subtotal" in lowered


def _looks_like_note(category: str) -> bool:
    lowered = category.lower()
    return lowered.startswith("note") or lowered.startswith("source")


def _get_row_category(worksheet, row: int) -> str:
    """Return the category text for *row*.

    Checks columns 1, 2, and 3 in order, returning the first non-empty value.
    This handles workbooks that use an indented multi-column layout where the
    category label appears in column 2 or 3 with earlier columns left blank.
    """
    for col in (1, 2, 3):
        cat = _to_text(worksheet.cell(row=row, column=col).value)
        if cat:
            return cat
    return ""


def _extract_unit_declaration(text: str) -> str | None:
    """Return the normalized unit string declared by *text*, or None.

    Detects three forms of unit declaration:

    1. A line whose primary content is a unit phrase, e.g.
       ``"Billions of dollars, by fiscal year"`` -> ``"Billions of dollars"``.
    2. A unit phrase inside a trailing parenthetical, e.g.
       ``"Federal Benefit Payments by Eligibility Category (Billions of dollars)"``
       -> ``"Billions of dollars"``.
    3. The standalone word ``"Dollars"`` (optionally with a footnote letter
       suffix), e.g. ``"Average Federal Spending per Enrollee (Dollars)c"``
       -> ``"Dollars"``.

    The returned string preserves the matched phrase's capitalization but
    strips footnote markers and surrounding punctuation so values are
    comparable downstream.
    """
    if not text:
        return None
    stripped = text.strip()
    if not stripped:
        return None
    # Form 3: standalone "Dollars" (line or paren); test first because it's
    # narrower than the generic phrase matcher.
    if DOLLARS_LINE_RE.match(stripped):
        return "Dollars"
    paren_dollars = DOLLARS_PAREN_RE.search(stripped)
    if paren_dollars:
        return "Dollars"
    # Form 2: unit phrase inside trailing parenthetical (e.g. section header).
    paren_match = UNIT_PAREN_RE.search(stripped)
    if paren_match:
        return paren_match.group(1).strip().rstrip(",").strip()
    # Form 1: line whose primary content is a unit phrase.
    line_match = UNIT_LINE_RE.match(stripped)
    if line_match:
        return line_match.group(1).strip().rstrip(",").strip()
    return None


def _normalize_unit_string(s: str) -> str:
    """Normalize a raw unit label (e.g. from the parse-plan YAML) to a consistent
    unit string.

    Handles embedded phrases such as ``'By Fiscal Year, Billions of Dollars'``
    or ``'Outlays in Millions of Dollars, by Fiscal Year'`` by finding the unit
    phrase anywhere in the string.  Returns the phrase with first-letter
    capitalisation and the remainder lower-cased for consistency across vintages.
    """
    if not s:
        return s
    phrase = _extract_unit_declaration(s)
    if phrase is None:
        m = re.search(rf"({_UNIT_PHRASE})", s, re.IGNORECASE)
        phrase = m.group(1).strip() if m else None
    if phrase:
        return phrase[:1].upper() + phrase[1:].lower()
    return s.strip()


def _build_unit_map(worksheet, plan: SheetPlan, year_columns: list[int]) -> dict[int, str]:
    """Return a map of *row* -> unit string in effect for that row.

    Pre-scans the sheet row-by-row, updating the running unit whenever a row's
    column-A text contains a unit declaration (see :func:`_extract_unit_declaration`).
    Rows that contain numeric data in any year column are NOT treated as unit
    declarations even if their text matches a unit pattern, because legitimate
    data rows occasionally include unit-like words.

    The default starting unit is ``plan.unit`` (from the parse plan), so sheets
    with no mid-sheet unit changes preserve their existing behavior.
    """
    row_unit: dict[int, str] = {}
    current = plan.unit
    for row in range(1, worksheet.max_row + 1):
        text_a = _to_text(worksheet.cell(row=row, column=1).value)
        if text_a:
            has_numeric_data = any(
                _parse_number(worksheet.cell(row=row, column=col).value) is not None
                for col in year_columns
            )
            if not has_numeric_data:
                detected = _extract_unit_declaration(text_a)
                if detected:
                    current = detected
        row_unit[row] = current
    return row_unit


def _extract_years(worksheet, plan: SheetPlan) -> dict[int, int]:
    """Return a mapping of column index → fiscal year found in that column's header.

    Scans up to ``MAX_YEAR_SCAN_ROWS`` rows (or the declared ``header_end_row``,
    whichever is larger) so that sheets whose year row falls below the declared
    header boundary are still handled correctly.

    * ``datetime`` / ``date`` cell values are skipped (publication timestamps).
    * Integer/float values are matched directly against the plausible-year range.
    * String values are matched with ``YEAR_LABEL_RE``, which requires the year
      to be the primary cell content.  This prevents publication-date strings
      like "February 2021" from being mis-classified as fiscal-year labels.
    """
    years: dict[int, int] = {}
    max_scan = max(plan.header_end_row, min(MAX_YEAR_SCAN_ROWS, worksheet.max_row))
    using_inferred_columns = not plan.year_columns
    if plan.year_columns:
        columns_to_scan = plan.year_columns
    else:
        max_inferred_column = min(worksheet.max_column, MAX_INFERRED_YEAR_COLUMNS)
        columns_to_scan = list(range(1, max_inferred_column + 1))
    for column in columns_to_scan:
        for row in range(1, max_scan + 1):
            value = worksheet.cell(row=row, column=column).value
            # Skip datetime / date objects — publication timestamps, not year labels.
            if isinstance(value, (datetime.datetime, datetime.date)):
                continue
            # Integer or float: test the value directly.
            if isinstance(value, (int, float)):
                year = int(value)
                if PLAUSIBLE_YEAR_MIN <= year <= PLAUSIBLE_YEAR_MAX:
                    years[column] = year
                    break
                continue  # value out of range — check the next row
            # String: require the cell to be essentially just a year label.
            cell_text = _to_text(value)
            match = YEAR_LABEL_RE.match(cell_text)
            if match:
                year = int(match.group(1))
                if PLAUSIBLE_YEAR_MIN <= year <= PLAUSIBLE_YEAR_MAX:
                    years[column] = year
                    break  # stop as soon as a valid year label is found
        if using_inferred_columns and len(years) >= MAX_INFERRED_YEAR_COLUMNS:
            break
    return years


def _infer_first_data_row(worksheet, plan: SheetPlan, year_columns: list[int] | None = None) -> int | None:
    active_year_columns = year_columns if year_columns is not None else plan.year_columns
    if not active_year_columns:
        return None
    for row in range(plan.header_end_row + 1, worksheet.max_row + 1):
        category = _get_row_category(worksheet, row)
        if not category:
            continue
        if _looks_like_note(category):
            continue
        parsed_values = [_parse_number(worksheet.cell(row=row, column=col).value) for col in active_year_columns]
        has_value = any(value is not None for value in parsed_values)
        if has_value:
            return row
    return None


def _infer_program_name(filename: str) -> str:
    """Infer a display program name from CBO workbook filenames.

    Typical filenames follow `<id>-<year>-<month>-<program>.xlsx`, e.g.
    `51293-2024-06-childnutrition.xlsx`. If this shape is not present, the
    entire stem is used.
    """
    stem = Path(filename).stem.replace("_", "-")
    parts = stem.split("-")
    if len(parts) >= 4:
        name = "-".join(parts[3:])
    else:
        name = stem
    return " ".join(token for token in re.split(r"[-\s]+", name) if token).title()


def _find_sheet(sheetnames: list[str], target: str) -> str | None:
    """Return the actual sheet name from *sheetnames* that matches *target*.

    Tries an exact match first, then falls back to a whitespace-stripped
    comparison so that sheet names with incidental trailing spaces (a common
    Excel authoring artefact) are handled transparently.
    """
    if target in sheetnames:
        return target
    target_stripped = target.strip()
    for name in sheetnames:
        if name.strip() == target_stripped:
            return name
    return None


def _read_plan(path: Path) -> list[SheetPlan]:
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    plans: list[SheetPlan] = []
    for workbook_entry in payload.get("workbooks", []):
        workbook_name = workbook_entry.get("workbook")
        for sheet_entry in workbook_entry.get("sheets", []):
            plans.append(
                SheetPlan(
                    workbook=workbook_name,
                    sheet=sheet_entry.get("sheet"),
                    include=bool(sheet_entry.get("include")),
                    output_dataset=sheet_entry.get("output_dataset", ""),
                    header_end_row=_header_end_row(str(sheet_entry.get("header_rows", "1-1"))),
                    first_data_row=sheet_entry.get("first_data_row"),
                    year_columns=[int(column) for column in sheet_entry.get("year_columns", [])],
                    unit=_normalize_unit_string(str(sheet_entry.get("unit", "")).strip()),
                    verification_exempt=bool(sheet_entry.get("verification_exempt", False)),
                )
            )
    return plans


def _in_slice(plan: SheetPlan, slice_name: str) -> bool:
    if slice_name == "all":
        return True
    dataset = plan.output_dataset.lower()
    if slice_name == "remaining-programs":
        return not any(kw in dataset for kw in HEALTH_KEYWORDS) and not any(
            kw in dataset for kw in INCOME_SECURITY_KEYWORDS
        )
    keywords = SLICE_KEYWORDS.get(slice_name, ())
    return any(keyword in dataset for keyword in keywords)


def _write_dataset(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _append_error(errors: list[str], workbook: str, sheet: str, reason: str) -> None:
    errors.append(f"{workbook}\t{sheet}\t{reason}")


def run_transform(
    parse_plan_path: Path = Path("config/workbook_parse_plan.yaml"),
    input_dir: Path = Path("data/raw"),
    output_dir: Path = Path("data/processed"),
    slice_name: str = "health",
) -> int:
    if slice_name not in SLICE_CHOICES:
        raise ValueError(f"Unsupported slice: {slice_name}")
    plans = [plan for plan in _read_plan(parse_plan_path) if plan.include and _in_slice(plan, slice_name) and not plan.verification_exempt]
    records_by_dataset: dict[str, list[dict]] = defaultdict(list)
    errors: list[str] = []

    for plan in plans:
        workbook_path = input_dir / plan.workbook
        if not workbook_path.exists():
            _append_error(errors, plan.workbook, plan.sheet, "workbook not found")
            continue
        workbook = load_workbook(workbook_path, data_only=True)
        try:
            actual_sheet = _find_sheet(workbook.sheetnames, plan.sheet)
            if actual_sheet is None:
                _append_error(errors, plan.workbook, plan.sheet, "sheet not found")
                continue
            worksheet = workbook[actual_sheet]
            years = _extract_years(worksheet, plan)
            if not years:
                _append_error(errors, plan.workbook, plan.sheet, "no fiscal years inferred")
                continue
            active_year_columns = sorted(years)
            first_data_row = plan.first_data_row or _infer_first_data_row(worksheet, plan, active_year_columns)
            if first_data_row is None:
                _append_error(errors, plan.workbook, plan.sheet, "could not infer first data row")
                continue

            # Single combined pass: track running unit AND collect data rows together.
            # Starting from row 1 (not first_data_row) ensures unit declarations in
            # header sections (e.g. row 11 in Medicaid) are seen before data rows.
            current_unit = plan.unit
            rows_written = 0
            program_name = _infer_program_name(plan.workbook)
            for row in range(1, worksheet.max_row + 1):
                text_a = _to_text(worksheet.cell(row=row, column=1).value)
                # Check for a unit-declaration row: has col-A text, no numeric data.
                if text_a:
                    has_numeric = any(
                        _parse_number(worksheet.cell(row=row, column=col).value) is not None
                        for col in active_year_columns
                    )
                    if not has_numeric:
                        detected = _extract_unit_declaration(text_a)
                        if detected:
                            current_unit = _normalize_unit_string(detected)
                        if row < first_data_row:
                            continue
                        # Non-numeric header row in data section — skip as data.
                        continue
                    # Row has numeric data — emit if past header.
                    if row < first_data_row:
                        continue
                else:
                    if row < first_data_row:
                        continue
                    # Blank col-A row in data section — check cols 2-3 for category.
                    has_numeric = any(
                        _parse_number(worksheet.cell(row=row, column=col).value) is not None
                        for col in active_year_columns
                    )
                    if not has_numeric:
                        continue
                category = _get_row_category(worksheet, row)
                if not category:
                    continue
                if _looks_like_note(category):
                    continue
                for column in active_year_columns:
                    year = years.get(column)
                    if year is None:
                        continue
                    value = _parse_number(worksheet.cell(row=row, column=column).value)
                    if value is None:
                        continue
                    records_by_dataset[plan.output_dataset].append(
                        {
                            "program": program_name,
                            "category": category,
                            "fiscal_year": year,
                            "value": value,
                            "unit": current_unit or "Unknown",
                            "source_file": plan.workbook,
                            "source_sheet": plan.sheet,
                            "is_total": str(_is_total(category)).lower(),
                        }
                    )
                    rows_written += 1
            if rows_written == 0:
                _append_error(errors, plan.workbook, plan.sheet, "no data rows parsed")
        finally:
            workbook.close()

    output_dir.mkdir(parents=True, exist_ok=True)
    for dataset, rows in records_by_dataset.items():
        _write_dataset(output_dir / f"{dataset}.csv", rows)

    error_path = output_dir / "parse_errors.log"
    error_body = "\n".join(errors)
    if errors:
        error_body += "\n"
    error_path.write_text(error_body, encoding="utf-8")
    print(
        f"Transform complete. slice={slice_name}, datasets={len(records_by_dataset)}, "
        f"rows={sum(len(rows) for rows in records_by_dataset.values())}, errors={len(errors)}"
    )
    return 1 if errors else 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Transform CBO baseline workbooks into tidy CSV datasets.")
    parser.add_argument("--parse-plan", type=Path, default=Path("config/workbook_parse_plan.yaml"))
    parser.add_argument("--input-dir", type=Path, default=Path("data/raw"))
    parser.add_argument("--output-dir", type=Path, default=Path("data/processed"))
    parser.add_argument("--slice", dest="slice_name", choices=SLICE_CHOICES, default="health")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    return run_transform(
        parse_plan_path=args.parse_plan,
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        slice_name=args.slice_name,
    )


if __name__ == "__main__":
    raise SystemExit(main())
