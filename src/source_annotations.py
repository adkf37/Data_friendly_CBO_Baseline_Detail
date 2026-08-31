from __future__ import annotations

"""Extract source row-label footnotes from Excel rich text.

CBO workbooks store footnote markers as superscript-formatted rich-text runs.
This module deliberately relies on that formatting instead of guessing from
ordinary trailing letters.  A marker is paired with the nearest following
definition carrying the same letter, which supports worksheets that restart
their footnote alphabet for each table.
"""

import re
from collections import defaultdict
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText


FOOTNOTE_START_RE = re.compile(r"^\s*([a-z])\s*[.)\]:-]\s*(.*?)\s*$", re.IGNORECASE)
MARKER_RUN_RE = re.compile(r"[a-z](?:\s*[,;]\s*[a-z])*|[a-z]+", re.IGNORECASE)
MISSING_FOOTNOTE_TEMPLATE = (
    "The source workbook includes superscript marker '{marker}' but does not "
    "provide a corresponding footnote definition on this worksheet."
)
KNOWN_MISSING_FOOTNOTE_DEFINITIONS = {
    (
        "51304-2026-02-pellgrant.xlsx",
        "T4_Pell_02-2026",
        30,
        1,
        "d",
    ),
}


@dataclass(frozen=True)
class SuperscriptLabel:
    row: int
    column: int
    raw_label: str
    base_label: str
    markers: tuple[str, ...]


@dataclass(frozen=True)
class FootnoteDefinition:
    row: int
    marker: str
    text: str


@dataclass(frozen=True)
class ResolvedAnnotation:
    label_row: int
    label_column: int
    raw_label: str
    base_label: str
    marker: str
    note_row: int | None
    variable_note: str


@dataclass(frozen=True)
class ObservationContext:
    source_file: str
    source_sheet: str
    source_row: int
    source_column: int
    category_path: str


@dataclass(frozen=True)
class VariableNote:
    source_file: str
    source_sheet: str
    label_row: int
    label_column: int
    category_path: str
    source_label: str
    marker: str
    variable_note: str


@dataclass
class AnnotationCatalog:
    by_source: dict[tuple[str, str], tuple[ResolvedAnnotation, ...]]
    missing_files: tuple[str, ...] = ()
    missing_sheets: tuple[tuple[str, str], ...] = ()

    @property
    def marker_references(self) -> int:
        return sum(len(items) for items in self.by_source.values())

    @property
    def resolved_marker_references(self) -> int:
        return sum(
            1
            for items in self.by_source.values()
            for item in items
            if item.variable_note
        )


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def rich_text_to_plain_text(value: object) -> str:
    """Return a whitespace-normalized string for plain or rich Excel text."""

    if value is None:
        return ""
    if isinstance(value, CellRichText):
        parts: list[str] = []
        for piece in value:
            if isinstance(piece, str):
                parts.append(piece)
            else:
                parts.append(getattr(piece, "text", str(piece)))
        return _normalize_text("".join(parts))
    return _normalize_text(value)


def _markers_from_superscript(text: str) -> list[str]:
    """Parse letter markers from one superscript run.

    In addition to the single or contiguous letters used by the reference CBO
    implementation, source workbooks contain runs such as ``a,b`` and
    ``g,h,i``.  Punctuation-only runs and superscripted numbers are not treated
    as footnote markers.
    """

    cleaned = re.sub(r"\s+", "", text).lower()
    if not cleaned or not MARKER_RUN_RE.fullmatch(cleaned):
        return []
    return re.findall(r"[a-z]", cleaned)


def extract_superscript_label(cell) -> SuperscriptLabel | None:
    """Extract a base label and actual superscript letter markers from *cell*."""

    value = cell.value
    if not isinstance(value, CellRichText):
        return None

    base_parts: list[str] = []
    markers: list[str] = []
    for piece in value:
        if isinstance(piece, str):
            base_parts.append(piece)
            continue

        text = getattr(piece, "text", "")
        font = getattr(piece, "font", None)
        if getattr(font, "vertAlign", None) == "superscript":
            parsed = _markers_from_superscript(text)
            if parsed:
                markers.extend(parsed)
            else:
                # Superscript formatting is sometimes applied to alignment-only
                # whitespace.  Preserve any non-marker content in the base label.
                base_parts.append(text)
        else:
            base_parts.append(text)

    base_label = _normalize_text("".join(base_parts))
    raw_label = rich_text_to_plain_text(value)
    markers = list(dict.fromkeys(markers))
    if len(base_label) < 4 or not markers:
        return None
    return SuperscriptLabel(
        row=cell.row,
        column=cell.column,
        raw_label=raw_label,
        base_label=base_label,
        markers=tuple(markers),
    )


def _worksheet_cells(worksheet) -> Iterable:
    """Iterate instantiated cells without expanding styled empty ranges."""

    cells = getattr(worksheet, "_cells", None)
    if isinstance(cells, dict):
        return cells.values()
    return (cell for row in worksheet.iter_rows() for cell in row)


def _sheet_content(worksheet) -> tuple[list[SuperscriptLabel], dict[int, list[tuple[int, str]]], set[int]]:
    labels: list[SuperscriptLabel] = []
    text_by_row: dict[int, list[tuple[int, str]]] = defaultdict(list)
    numeric_rows: set[int] = set()

    for cell in _worksheet_cells(worksheet):
        text = rich_text_to_plain_text(cell.value)
        if text:
            text_by_row[cell.row].append((cell.column, text))
        if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
            numeric_rows.add(cell.row)
        label = extract_superscript_label(cell)
        if label is not None:
            labels.append(label)

    for cells in text_by_row.values():
        cells.sort()
    labels.sort(key=lambda item: (item.row, item.column))
    return labels, text_by_row, numeric_rows


def _extract_footnote_definitions(
    text_by_row: dict[int, list[tuple[int, str]]],
    numeric_rows: set[int],
    valid_markers: set[str],
) -> list[FootnoteDefinition]:
    """Extract every source footnote block, including wrapped continuation rows."""

    pending_row: int | None = None
    pending_marker = ""
    pending_parts: list[str] = []
    definitions: list[FootnoteDefinition] = []
    previous_row: int | None = None

    def flush() -> None:
        nonlocal pending_row, pending_marker, pending_parts
        if pending_row is not None and pending_parts:
            definitions.append(
                FootnoteDefinition(
                    row=pending_row,
                    marker=pending_marker,
                    text=_normalize_text(" ".join(pending_parts)),
                )
            )
        pending_row = None
        pending_marker = ""
        pending_parts = []

    for row in sorted(text_by_row):
        if previous_row is not None and row > previous_row + 1:
            flush()
        previous_row = row
        row_cells = text_by_row[row]
        first_text = row_cells[0][1]
        row_text = " ".join(text for _, text in row_cells).strip()
        start = FOOTNOTE_START_RE.match(first_text)
        if start and start.group(1).lower() in valid_markers:
            flush()
            marker = start.group(1).lower()
            first_note_part = start.group(2).strip()
            remaining = " ".join(text for _, text in row_cells[1:]).strip()
            pending_row = row
            pending_marker = marker
            pending_parts = [part for part in (first_note_part, remaining) if part]
            continue

        if pending_row is None:
            continue
        if row in numeric_rows or row_text.casefold().startswith("source:"):
            flush()
            continue
        pending_parts.append(row_text)

    flush()
    return definitions


def extract_worksheet_annotations(worksheet) -> tuple[ResolvedAnnotation, ...]:
    """Resolve every rich-text marker on a worksheet to its source note text."""

    labels, text_by_row, numeric_rows = _sheet_content(worksheet)
    if not labels:
        return ()
    valid_markers = {marker for label in labels for marker in label.markers}
    definitions = _extract_footnote_definitions(text_by_row, numeric_rows, valid_markers)
    definitions_by_marker: dict[str, list[FootnoteDefinition]] = defaultdict(list)
    for definition in definitions:
        definitions_by_marker[definition.marker].append(definition)

    resolved: list[ResolvedAnnotation] = []
    for label in labels:
        for marker in label.markers:
            following = [
                definition
                for definition in definitions_by_marker.get(marker, [])
                if definition.row > label.row
            ]
            definition = min(following, key=lambda item: item.row) if following else None
            resolved.append(
                ResolvedAnnotation(
                    label_row=label.row,
                    label_column=label.column,
                    raw_label=label.raw_label,
                    base_label=label.base_label,
                    marker=marker,
                    note_row=definition.row if definition else None,
                    variable_note=definition.text if definition else "",
                )
            )
    return tuple(resolved)


def _resolve_sheet_name(sheetnames: list[str], requested: str) -> str | None:
    if requested in sheetnames:
        return requested
    requested_stripped = requested.strip()
    for sheet_name in sheetnames:
        if sheet_name.strip() == requested_stripped:
            return sheet_name
    return None


def load_annotation_catalog(
    raw_dir: Path,
    requested_sources: dict[str, set[str]],
) -> AnnotationCatalog:
    """Load rich-text annotations for the source workbook/sheet pairs requested."""

    by_source: dict[tuple[str, str], tuple[ResolvedAnnotation, ...]] = {}
    missing_files: list[str] = []
    missing_sheets: list[tuple[str, str]] = []

    for source_file, requested_sheets in sorted(requested_sources.items()):
        workbook_path = raw_dir / source_file
        if not workbook_path.exists():
            missing_files.append(source_file)
            continue
        workbook = load_workbook(
            workbook_path,
            data_only=True,
            read_only=False,
            rich_text=True,
        )
        try:
            for requested_sheet in sorted(requested_sheets):
                actual_sheet = _resolve_sheet_name(workbook.sheetnames, requested_sheet)
                if actual_sheet is None:
                    missing_sheets.append((source_file, requested_sheet))
                    continue
                annotations = extract_worksheet_annotations(workbook[actual_sheet])
                by_source[(source_file, requested_sheet)] = tuple(
                    replace(
                        item,
                        variable_note=MISSING_FOOTNOTE_TEMPLATE.format(marker=item.marker),
                    )
                    if not item.variable_note
                    and (
                        source_file,
                        requested_sheet,
                        item.label_row,
                        item.label_column,
                        item.marker,
                    )
                    in KNOWN_MISSING_FOOTNOTE_DEFINITIONS
                    else item
                    for item in annotations
                )
        finally:
            workbook.close()

    return AnnotationCatalog(
        by_source=by_source,
        missing_files=tuple(missing_files),
        missing_sheets=tuple(missing_sheets),
    )


def _annotations_for_context(
    context: ObservationContext,
    annotations: tuple[ResolvedAnnotation, ...],
) -> list[ResolvedAnnotation]:
    components = [part.strip() for part in context.category_path.split(" / ") if part.strip()]
    component_keys = {component.casefold() for component in components}
    selected: list[ResolvedAnnotation] = []

    # Prefer the source-faithful label, including its visible marker. A base
    # label is accepted only on the exact source row; a broader base match would
    # incorrectly spread a note to later rows that reuse the same label (for
    # example repeated "Annual Subsidy Costs").
    for component_key in component_keys:
        candidates = [
            item
            for item in annotations
            if item.label_row <= context.source_row and item.raw_label.casefold() == component_key
        ]
        if not candidates:
            candidates = [
                item
                for item in annotations
                if item.label_row == context.source_row
                and item.base_label.casefold() == component_key
            ]
        if not candidates:
            continue
        latest_row = max(item.label_row for item in candidates)
        latest_column = max(item.label_column for item in candidates if item.label_row == latest_row)
        selected.extend(
            item
            for item in candidates
            if item.label_row == latest_row and item.label_column == latest_column
        )
    return selected


def match_variable_notes(
    contexts: Iterable[ObservationContext],
    catalog: AnnotationCatalog,
) -> list[VariableNote]:
    """Bind resolved source annotations to affected dataset category paths."""

    notes: dict[tuple, VariableNote] = {}
    contexts_by_source: dict[tuple[str, str], list[ObservationContext]] = defaultdict(list)
    for context in contexts:
        contexts_by_source[(context.source_file, context.source_sheet)].append(context)

    def add(context: ObservationContext, item: ResolvedAnnotation) -> None:
        note = VariableNote(
            source_file=context.source_file,
            source_sheet=context.source_sheet,
            label_row=item.label_row,
            label_column=item.label_column,
            category_path=context.category_path,
            source_label=item.raw_label,
            marker=item.marker,
            variable_note=item.variable_note,
        )
        key = (
            note.source_file,
            note.source_sheet,
            note.label_row,
            note.label_column,
            note.category_path,
            note.source_label,
            note.marker,
            note.variable_note,
        )
        notes[key] = note

    for source, source_contexts in contexts_by_source.items():
        annotations = catalog.by_source.get(source, ())
        matched_references: set[tuple[int, int, str]] = set()
        for context in source_contexts:
            for item in _annotations_for_context(context, annotations):
                add(context, item)
                matched_references.add((item.label_row, item.label_column, item.marker))

        # Some source tables put a superscript note on a spanning header that
        # is not represented verbatim in category_path. Bind those headers by
        # their row/column region. The five-row proximity requirement prevents
        # an orphaned label from an unparsed table being applied to a later,
        # unrelated table.
        label_cells = sorted({(item.label_row, item.label_column) for item in annotations})
        for item in annotations:
            reference_key = (item.label_row, item.label_column, item.marker)
            if reference_key in matched_references:
                continue
            next_row = min(
                (
                    row
                    for row, column in label_cells
                    if column == item.label_column and row > item.label_row
                ),
                default=item.note_row or 10**9,
            )
            if item.note_row is not None:
                next_row = min(next_row, item.note_row)
            next_column = min(
                (
                    column
                    for row, column in label_cells
                    if row == item.label_row and column > item.label_column
                ),
                default=10**9,
            )
            candidates = [
                context
                for context in source_contexts
                if item.label_row <= context.source_row < next_row
                and item.label_column <= context.source_column < next_column
            ]
            if not candidates:
                continue
            first_observation_row = min(context.source_row for context in candidates)
            if first_observation_row - item.label_row > 5:
                continue
            for context in candidates:
                add(context, item)
            matched_references.add(reference_key)

        represented = {
            (note.source_label, note.marker, note.variable_note)
            for note in notes.values()
            if (note.source_file, note.source_sheet) == source
        }
        for item in annotations:
            identity = (item.raw_label, item.marker, item.variable_note)
            if identity in represented:
                continue
            source_only = VariableNote(
                source_file=source[0],
                source_sheet=source[1],
                label_row=item.label_row,
                label_column=item.label_column,
                category_path="",
                source_label=item.raw_label,
                marker=item.marker,
                variable_note=item.variable_note,
            )
            key = (
                source_only.source_file,
                source_only.source_sheet,
                source_only.label_row,
                source_only.label_column,
                source_only.category_path,
                source_only.source_label,
                source_only.marker,
                source_only.variable_note,
            )
            notes[key] = source_only
            represented.add(identity)
    return sorted(
        notes.values(),
        key=lambda item: (
            item.source_file,
            item.source_sheet,
            item.label_row,
            item.label_column,
            item.category_path,
            item.marker,
        ),
    )
