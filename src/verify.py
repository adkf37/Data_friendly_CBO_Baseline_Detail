from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

if sys.path:
    script_dir = Path(__file__).resolve().parent
    if Path(sys.path[0]).resolve() == script_dir:
        sys.path[0] = str(script_dir.parent)

import yaml
from openpyxl import load_workbook

from src import transform

RELATIVE_TOLERANCE = 0.0001  # 0.01%
VALID_PERIOD_TYPES = {
    "fiscal_year",
    "calendar_year",
    "award_year",
    "school_year",
    "cumulative_fiscal_years",
    "unmapped",
}


@dataclass(frozen=True)
class VerificationPlan:
    workbook: str
    sheet: str
    include: bool
    output_dataset: str
    verification_target: str
    header_end_row: int
    first_data_row: int | None
    year_columns: list[int]
    unit: str
    verification_include_totals: bool
    verification_exempt: bool
    verification_exempt_reason: str


@dataclass(frozen=True)
class YearComparison:
    fiscal_year: int
    source_total: float
    processed_total: float
    absolute_diff: float
    relative_diff: float | None
    passed: bool


@dataclass(frozen=True)
class VerificationResult:
    plan: VerificationPlan
    status: str
    mode: str
    absolute_tolerance: float
    relative_tolerance: float
    comparisons: list[YearComparison]
    source_cells: int
    processed_rows: int
    matched_cells: int
    missing_cells: int
    duplicate_cells: int
    unexpected_cells: int
    invalid_rows: int
    notes: list[str]
    error_details: list[str]


def _read_plan(parse_plan_path: Path) -> list[VerificationPlan]:
    payload = yaml.safe_load(parse_plan_path.read_text(encoding="utf-8")) or {}
    plans: list[VerificationPlan] = []
    for workbook_entry in payload.get("workbooks", []):
        workbook_name = workbook_entry.get("workbook", "")
        for sheet_entry in workbook_entry.get("sheets", []):
            plans.append(
                VerificationPlan(
                    workbook=workbook_name,
                    sheet=sheet_entry.get("sheet", ""),
                    include=bool(sheet_entry.get("include")),
                    output_dataset=sheet_entry.get("output_dataset", ""),
                    verification_target=sheet_entry.get("verification_target", ""),
                    header_end_row=transform._header_end_row(str(sheet_entry.get("header_rows", "1-1"))),
                    first_data_row=sheet_entry.get("first_data_row"),
                    year_columns=[int(column) for column in sheet_entry.get("year_columns", [])],
                    unit=str(sheet_entry.get("unit", "")).strip(),
                    verification_include_totals=bool(sheet_entry.get("verification_include_totals", False)),
                    verification_exempt=bool(sheet_entry.get("verification_exempt", False)),
                    verification_exempt_reason=str(sheet_entry.get("verification_exempt_reason", "")).strip(),
                )
            )
    return plans


def _sheet_plan(plan: VerificationPlan) -> transform.SheetPlan:
    return transform.SheetPlan(
        workbook=plan.workbook,
        sheet=plan.sheet,
        include=plan.include,
        output_dataset=plan.output_dataset,
        header_end_row=plan.header_end_row,
        first_data_row=plan.first_data_row,
        year_columns=plan.year_columns,
        unit=transform._normalize_unit_string(plan.unit),
        verification_exempt=plan.verification_exempt,
    )


def _absolute_tolerance_for_unit(unit: str) -> float:
    lowered = unit.lower()
    if "billion" in lowered:
        return 0.001
    if "million" in lowered:
        return 0.01
    if "thousand" in lowered:
        return 1.0
    if "percent" in lowered or "share" in lowered or "rate" in lowered:
        return 0.001
    return 0.01


def _target_name(plan: VerificationPlan) -> str:
    return plan.verification_target or f"{plan.output_dataset}:{plan.sheet}"


def _read_dataset(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    if not path.exists():
        return [], []
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader), list(reader.fieldnames or [])


def _expected_source_cells(worksheet, plan: VerificationPlan) -> dict[tuple[int, int], transform.PeriodColumn | None]:
    """Inventory numeric source measures independently of processed totals."""

    sheet_plan = _sheet_plan(plan)
    blocks = transform._period_blocks(worksheet, sheet_plan)
    expected: dict[tuple[int, int], transform.PeriodColumn | None] = {}
    if blocks:
        blocks_by_row: dict[int, list[transform.PeriodColumn]] = {}
        for block in blocks:
            blocks_by_row.setdefault(block.header_row, []).extend(block.periods)
        current_periods: list[transform.PeriodColumn] = []
        first_data_row = plan.first_data_row or min(block.header_row for block in blocks) + 1
        for row in range(1, worksheet.max_row + 1):
            if row in blocks_by_row:
                by_column = {period.column: period for period in blocks_by_row[row]}
                current_periods = [by_column[column] for column in sorted(by_column)]
                continue
            if row < first_data_row or not current_periods:
                continue
            period_start = min(period.column for period in current_periods)
            text_cells = transform._row_text_cells(worksheet, row, period_start)
            if any(transform._looks_like_note(text) for _, text in text_cells):
                continue
            for period in current_periods:
                if transform._parse_number(worksheet.cell(row, period.column).value) is not None:
                    expected[(row, period.column)] = period
        return expected

    first_data_row = plan.first_data_row or plan.header_end_row + 1
    max_column = min(worksheet.max_column, 200)
    for row in range(first_data_row, worksheet.max_row + 1):
        numeric_columns = [
            column
            for column in range(1, max_column + 1)
            if transform._parse_number(worksheet.cell(row, column).value) is not None
        ]
        leading_year_column = next(
            (
                column
                for column in numeric_columns
                if column <= 3 and transform._exact_year(worksheet.cell(row, column).value) is not None
            ),
            None,
        )
        for column in numeric_columns:
            if column != leading_year_column:
                expected[(row, column)] = None
    return expected


def _integer_field(row: dict[str, str], field: str) -> int | None:
    text = (row.get(field) or "").strip()
    if not text:
        return None
    try:
        value = float(text)
    except ValueError:
        return None
    return int(value) if value.is_integer() else None


def _period_errors(row: dict[str, str], expected: transform.PeriodColumn | None) -> list[str]:
    errors: list[str] = []
    period_type = (row.get("period_type") or "").strip()
    start_year = _integer_field(row, "period_start_year")
    end_year = _integer_field(row, "period_end_year")
    fiscal_year = _integer_field(row, "fiscal_year")
    category_path = (row.get("category_path") or "").strip()
    if period_type not in VALID_PERIOD_TYPES:
        return [f"invalid period_type {period_type!r}"]
    if period_type == "fiscal_year":
        if fiscal_year is None or fiscal_year != end_year or start_year != end_year:
            errors.append("fiscal-year fields are inconsistent")
    elif fiscal_year is not None:
        errors.append("non-fiscal period populated fiscal_year")
    if period_type == "cumulative_fiscal_years" and (
        start_year is None or end_year is None or start_year >= end_year
    ):
        errors.append("invalid cumulative-year range")
    if expected is not None:
        expected_type = transform._period_type_for_row(expected.period_type, category_path)
        if period_type != expected_type:
            errors.append(f"period_type {period_type!r} != source-derived {expected_type!r}")
        if start_year != expected.start_year or end_year != expected.end_year:
            errors.append("period bounds do not match source header")
        if (row.get("period_label") or "").strip() != expected.label:
            errors.append("period_label does not match source header")
    elif period_type not in {"unmapped", "fiscal_year", "calendar_year", "award_year", "school_year"}:
        errors.append("generic-table period is not supported")
    return errors


def _direct_result(
    plan: VerificationPlan,
    worksheet,
    dataset_rows: list[dict[str, str]],
    fieldnames: list[str],
) -> VerificationResult:
    scoped_rows = [
        row
        for row in dataset_rows
        if row.get("source_file") == plan.workbook and row.get("source_sheet") == plan.sheet
    ]
    expected = _expected_source_cells(worksheet, plan)
    required = set(transform.OUTPUT_COLUMNS)
    errors: list[str] = []
    if not required.issubset(fieldnames):
        errors.append(f"processed schema missing: {', '.join(sorted(required - set(fieldnames)))}")

    coordinates: list[tuple[int, int]] = []
    invalid_rows = 0
    matched_cells = 0
    explicit_unit_by_row: dict[int, str | None] = {}
    for index, row in enumerate(scoped_rows, start=2):
        source_row = _integer_field(row, "source_row")
        source_column = _integer_field(row, "source_column")
        if source_row is None or source_column is None or source_row < 1 or source_column < 1:
            invalid_rows += 1
            errors.append(f"CSV row {index}: invalid source coordinate")
            continue
        coordinate = (source_row, source_column)
        coordinates.append(coordinate)
        source_value = transform._parse_number(worksheet.cell(source_row, source_column).value)
        try:
            processed_value = float((row.get("value") or "").strip())
        except ValueError:
            processed_value = None
        unit = (row.get("unit") or "").strip()
        tolerance = _absolute_tolerance_for_unit(unit)
        if source_value is None or processed_value is None or abs(source_value - processed_value) > tolerance:
            invalid_rows += 1
            errors.append(
                f"CSV row {index}: value {row.get('value')!r} does not match source R{source_row}C{source_column}={source_value!r}"
            )
        else:
            matched_cells += 1

        row_errors: list[str] = []
        if (row.get("program_id") or "").strip() != transform._program_id(plan.workbook):
            row_errors.append("program_id mismatch")
        if (row.get("program") or "").strip() != transform._canonical_program_name(plan.workbook):
            row_errors.append("program label mismatch")
        category = (row.get("category") or "").strip()
        category_path = (row.get("category_path") or "").strip()
        if not category or transform._parse_number(category) is not None:
            row_errors.append("category is blank or numeric")
        if not category_path:
            row_errors.append("category_path is blank")
        if source_row not in explicit_unit_by_row:
            explicit_unit_by_row[source_row] = transform._row_unit(
                transform._row_text_cells(worksheet, source_row)
            )
        explicit_unit = explicit_unit_by_row[source_row]
        if explicit_unit and transform._normalize_unit_string(unit) != explicit_unit:
            row_errors.append(f"unit {unit!r} contradicts source row unit {explicit_unit!r}")
        row_errors.extend(_period_errors(row, expected.get(coordinate)))
        if row_errors:
            invalid_rows += 1
            errors.append(f"CSV row {index}: " + "; ".join(row_errors))

    counts = Counter(coordinates)
    duplicate_cells = sum(count - 1 for count in counts.values() if count > 1)
    processed_coordinates = set(coordinates)
    missing = sorted(set(expected) - processed_coordinates)
    unexpected = sorted(processed_coordinates - set(expected))
    if duplicate_cells:
        errors.append(f"{duplicate_cells} duplicate source-cell references")
    if missing:
        errors.append(f"{len(missing)} numeric source cells omitted; first: {missing[:5]}")
    if unexpected:
        errors.append(f"{len(unexpected)} processed coordinates outside source measure scope; first: {unexpected[:5]}")
    if not scoped_rows:
        errors.append("no processed rows matched source_file/source_sheet scope")

    return VerificationResult(
        plan=plan,
        status="FAIL" if errors else "PASS",
        mode="direct source-cell",
        absolute_tolerance=_absolute_tolerance_for_unit(plan.unit),
        relative_tolerance=RELATIVE_TOLERANCE,
        comparisons=[],
        source_cells=len(expected),
        processed_rows=len(scoped_rows),
        matched_cells=matched_cells,
        missing_cells=len(missing),
        duplicate_cells=duplicate_cells,
        unexpected_cells=len(unexpected),
        invalid_rows=invalid_rows,
        notes=[],
        error_details=errors[:25],
    )


def _legacy_totals(plan: VerificationPlan, worksheet, rows: list[dict[str, str]]) -> VerificationResult:
    """Compatibility check for historical eight-column CSV fixtures."""

    sheet_plan = _sheet_plan(plan)
    years = transform._extract_years(worksheet, sheet_plan)
    first_data_row = plan.first_data_row or transform._infer_first_data_row(worksheet, sheet_plan, sorted(years))
    source_totals: dict[int, float] = {}
    if first_data_row is not None:
        for source_row in range(first_data_row, worksheet.max_row + 1):
            category = transform._get_row_category(worksheet, source_row)
            if not category or transform._looks_like_note(category):
                continue
            if not plan.verification_include_totals and transform._is_total(category):
                continue
            for column, year in years.items():
                value = transform._parse_number(worksheet.cell(source_row, column).value)
                if value is not None:
                    source_totals[year] = source_totals.get(year, 0.0) + value

    processed_totals: dict[int, float] = {}
    scoped = [row for row in rows if row.get("source_file") == plan.workbook and row.get("source_sheet") == plan.sheet]
    for row in scoped:
        if not plan.verification_include_totals and (row.get("is_total") or "").lower() == "true":
            continue
        try:
            year = int(float(row.get("fiscal_year") or ""))
            value = float(row.get("value") or "")
        except ValueError:
            continue
        processed_totals[year] = processed_totals.get(year, 0.0) + value

    comparisons: list[YearComparison] = []
    tolerance = _absolute_tolerance_for_unit(plan.unit)
    for year in sorted(set(source_totals) | set(processed_totals)):
        source = source_totals.get(year, 0.0)
        processed = processed_totals.get(year, 0.0)
        absolute = abs(source - processed)
        relative = None if source == 0 else absolute / abs(source)
        passed = absolute <= tolerance or (relative is not None and relative <= RELATIVE_TOLERANCE)
        comparisons.append(YearComparison(year, source, processed, absolute, relative, passed))
    failed = not comparisons or any(not item.passed for item in comparisons)
    notes = ["legacy aggregate mode: source-cell lineage columns are absent"]
    return VerificationResult(
        plan=plan,
        status="FAIL" if failed else "PASS",
        mode="legacy aggregate",
        absolute_tolerance=tolerance,
        relative_tolerance=RELATIVE_TOLERANCE,
        comparisons=comparisons,
        source_cells=0,
        processed_rows=len(scoped),
        matched_cells=0,
        missing_cells=0,
        duplicate_cells=0,
        unexpected_cells=0,
        invalid_rows=0,
        notes=notes,
        error_details=[] if not failed else ["one or more annual aggregates differ"],
    )


def _missing_result(plan: VerificationPlan, message: str) -> VerificationResult:
    return VerificationResult(
        plan=plan,
        status="FAIL",
        mode="direct source-cell",
        absolute_tolerance=_absolute_tolerance_for_unit(plan.unit),
        relative_tolerance=RELATIVE_TOLERANCE,
        comparisons=[],
        source_cells=0,
        processed_rows=0,
        matched_cells=0,
        missing_cells=0,
        duplicate_cells=0,
        unexpected_cells=0,
        invalid_rows=0,
        notes=[],
        error_details=[message],
    )


def _render_report(results: list[VerificationResult]) -> str:
    failures = sum(result.status == "FAIL" for result in results)
    passes = sum(result.status == "PASS" for result in results)
    direct = [result for result in results if result.mode == "direct source-cell"]
    lines = [
        "# Verification Report",
        "",
        "Generated by `python src/verify.py`.",
        "",
        "## Summary",
        "",
        f"- Verification targets: {len(results)}",
        f"- Pass: {passes}",
        "- Exempt failures: 0",
        f"- Non-exempt failures: {failures}",
        f"- Direct source cells checked: {sum(result.source_cells for result in direct):,}",
        f"- Direct values matched: {sum(result.matched_cells for result in direct):,}",
        f"- Legacy aggregate targets: {sum(result.mode == 'legacy aggregate' for result in results)}",
        "",
        "Former parse-plan exemptions are included and must pass the same checks as every other target.",
        "",
        "## Target Status",
        "",
        "| Target | Dataset | Source | Mode | Cells | Status | Notes |",
        "|---|---|---|---|---:|---|---|",
    ]
    for result in results:
        source = f"{result.plan.workbook} :: {result.plan.sheet}"
        note = "; ".join(result.error_details[:2] or result.notes) or "—"
        lines.append(
            f"| `{_target_name(result.plan)}` | `{result.plan.output_dataset}` | `{source}` | "
            f"{result.mode} | {result.source_cells} | **{result.status}** | {note} |"
        )

    lines.extend(["", "## Target Details", ""])
    for result in results:
        lines.extend(
            [
                f"### `{_target_name(result.plan)}`",
                "",
                f"- Dataset: `{result.plan.output_dataset}`",
                f"- Source: `{result.plan.workbook}` / `{result.plan.sheet}`",
                f"- Status: **{result.status}**",
                f"- Verification mode: {result.mode}",
            ]
        )
        if result.mode == "direct source-cell":
            lines.append(
                f"- Source cells: {result.source_cells}; processed rows: {result.processed_rows}; "
                f"matched values: {result.matched_cells}; missing: {result.missing_cells}; "
                f"duplicate: {result.duplicate_cells}; unexpected: {result.unexpected_cells}; "
                f"invalid rows: {result.invalid_rows}"
            )
        if result.notes:
            lines.append(f"- Notes: {'; '.join(result.notes)}")
        if result.error_details:
            lines.append(f"- Errors: {'; '.join(result.error_details)}")
        if result.comparisons:
            lines.extend(
                [
                    "",
                    "| Fiscal Year | Source Total | Processed Total | Absolute Diff | Relative Diff | Result |",
                    "|---|---:|---:|---:|---:|---|",
                ]
            )
            for comparison in result.comparisons:
                relative = "n/a" if comparison.relative_diff is None else f"{comparison.relative_diff * 100:.6f}%"
                lines.append(
                    f"| {comparison.fiscal_year} | {comparison.source_total:.6f} | "
                    f"{comparison.processed_total:.6f} | {comparison.absolute_diff:.6f} | {relative} | "
                    f"{'PASS' if comparison.passed else 'FAIL'} |"
                )
        lines.append("")
    return "\n".join(lines) + "\n"


def run_verification(
    parse_plan_path: Path = Path("config/workbook_parse_plan.yaml"),
    input_dir: Path = Path("data/raw"),
    processed_dir: Path = Path("data/processed"),
    report_path: Path = Path("docs/verification_report.md"),
) -> int:
    plans = [plan for plan in _read_plan(parse_plan_path) if plan.include]
    dataset_cache: dict[str, tuple[list[dict[str, str]], list[str]]] = {}
    results: list[VerificationResult] = []
    plans_by_workbook: dict[str, list[VerificationPlan]] = {}
    for plan in plans:
        plans_by_workbook.setdefault(plan.workbook, []).append(plan)

    for workbook_name, workbook_plans in plans_by_workbook.items():
        workbook_path = input_dir / workbook_name
        if not workbook_path.exists():
            results.extend(_missing_result(plan, f"workbook missing: {workbook_name}") for plan in workbook_plans)
            continue
        workbook = load_workbook(workbook_path, data_only=True)
        try:
            for plan in workbook_plans:
                actual_sheet = transform._find_sheet(workbook.sheetnames, plan.sheet)
                if actual_sheet is None:
                    results.append(_missing_result(plan, f"sheet missing: {plan.sheet}"))
                    continue
                if plan.output_dataset not in dataset_cache:
                    dataset_cache[plan.output_dataset] = _read_dataset(processed_dir / f"{plan.output_dataset}.csv")
                rows, fieldnames = dataset_cache[plan.output_dataset]
                if not fieldnames:
                    results.append(_missing_result(plan, f"processed CSV missing: {plan.output_dataset}.csv"))
                elif {"source_row", "source_column"}.issubset(fieldnames):
                    results.append(_direct_result(plan, workbook[actual_sheet], rows, fieldnames))
                else:
                    results.append(_legacy_totals(plan, workbook[actual_sheet], rows))
        finally:
            workbook.close()

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(_render_report(results), encoding="utf-8")
    failures = sum(result.status == "FAIL" for result in results)
    print(
        "Verification complete. "
        f"targets={len(results)}, pass={sum(result.status == 'PASS' for result in results)}, "
        f"failures={failures}, report={report_path}"
    )
    return 1 if failures else 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Verify processed CBO CSV rows against exact source cells.")
    parser.add_argument("--parse-plan", type=Path, default=Path("config/workbook_parse_plan.yaml"))
    parser.add_argument("--input-dir", type=Path, default=Path("data/raw"))
    parser.add_argument("--processed-dir", type=Path, default=Path("data/processed"))
    parser.add_argument("--report", type=Path, default=Path("docs/verification_report.md"))
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    return run_verification(args.parse_plan, args.input_dir, args.processed_dir, args.report)


if __name__ == "__main__":
    raise SystemExit(main())
