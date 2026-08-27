import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from src import generate_schemas


def _write_sample_csv(path: Path, *, has_totals: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "program,category,fiscal_year,value,unit,source_file,source_sheet,is_total,program_id,category_path,period_type,period_start_year,period_end_year,period_label,source_row,source_column\n"
        "SNAP,Estimated Budget Authority,2024,103764.0,Millions of dollars,51312-2024-06-snap.xlsx,SNAP_06-2024,false,51312,Estimated Budget Authority,fiscal_year,2024,2024,2024,10,4\n"
        + (
            "SNAP,Total Outlays,2024,104000.0,Millions of dollars,51312-2024-06-snap.xlsx,SNAP_06-2024,true,51312,Total Outlays,fiscal_year,2024,2024,2024,11,4\n"
            if has_totals
            else ""
        ),
        encoding="utf-8",
    )


class GenerateSchemasTests(unittest.TestCase):
    def test_generate_schemas_creates_one_file_per_csv(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"

            _write_sample_csv(processed / "snap_2024_06.csv")
            _write_sample_csv(processed / "medicaid_2024_06.csv")

            rc = generate_schemas.generate_schemas(processed_dir=processed, schemas_dir=schemas)

            self.assertEqual(0, rc)
            self.assertTrue((schemas / "snap_2024_06.md").exists())
            self.assertTrue((schemas / "medicaid_2024_06.md").exists())

    def test_generate_schemas_creates_readme_index(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"

            _write_sample_csv(processed / "snap_2024_06.csv")

            generate_schemas.generate_schemas(processed_dir=processed, schemas_dir=schemas)

            readme = schemas / "README.md"
            self.assertTrue(readme.exists())
            content = readme.read_text(encoding="utf-8")
            self.assertIn("snap_2024_06", content)
            self.assertIn("snap_2024_06.md", content)
            self.assertIn("Total datasets:", content)

    def test_schema_file_contains_required_sections(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"

            _write_sample_csv(processed / "snap_2024_06.csv")
            generate_schemas.generate_schemas(processed_dir=processed, schemas_dir=schemas)

            content = (schemas / "snap_2024_06.md").read_text(encoding="utf-8")
            self.assertIn("## Purpose", content)
            self.assertIn("## Provenance", content)
            self.assertIn("## Columns", content)
            self.assertIn("## is_total Interpretation", content)

    def test_schema_file_documents_all_output_columns(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"

            _write_sample_csv(processed / "snap_2024_06.csv")
            generate_schemas.generate_schemas(processed_dir=processed, schemas_dir=schemas)

            content = (schemas / "snap_2024_06.md").read_text(encoding="utf-8")
            for col_name in generate_schemas.COLUMN_META:
                col_name = col_name["name"]
                self.assertIn(col_name, content, f"Column '{col_name}' missing from schema")

    def test_schema_file_includes_provenance_from_csv(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"

            _write_sample_csv(processed / "snap_2024_06.csv")
            generate_schemas.generate_schemas(processed_dir=processed, schemas_dir=schemas)

            content = (schemas / "snap_2024_06.md").read_text(encoding="utf-8")
            self.assertIn("51312-2024-06-snap.xlsx", content)
            self.assertIn("SNAP_06-2024", content)

    def test_schema_file_warns_about_totals_when_present(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"

            _write_sample_csv(processed / "snap_2024_06.csv", has_totals=True)
            generate_schemas.generate_schemas(processed_dir=processed, schemas_dir=schemas)

            content = (schemas / "snap_2024_06.md").read_text(encoding="utf-8")
            self.assertIn("double-counting", content)

    def test_schema_includes_superscript_variable_note_from_raw_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"
            raw = root / "raw"
            raw.mkdir()
            _write_sample_csv(processed / "snap_2024_06.csv")

            workbook = Workbook()
            ws = workbook.active
            ws.title = "SNAP_06-2024"
            ws["A10"] = CellRichText(
                [
                    "Estimated Budget Authority",
                    TextBlock(InlineFont(vertAlign="superscript"), "a"),
                ]
            )
            ws["D10"] = 103764
            ws["A12"] = "a. Includes a source-specific adjustment."
            workbook.save(raw / "51312-2024-06-snap.xlsx")

            generate_schemas.generate_schemas(
                processed_dir=processed,
                schemas_dir=schemas,
                raw_dir=raw,
            )

            content = (schemas / "snap_2024_06.md").read_text(encoding="utf-8")
            self.assertIn("## Variable Notes", content)
            self.assertIn("`variable_note`", content)
            self.assertIn("Includes a source-specific adjustment.", content)
            self.assertIn("R10C1", content)

    def test_schema_generation_fails_audit_when_note_text_is_missing(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"
            raw = root / "raw"
            raw.mkdir()
            _write_sample_csv(processed / "snap_2024_06.csv")

            workbook = Workbook()
            ws = workbook.active
            ws.title = "SNAP_06-2024"
            ws["A10"] = CellRichText(
                [
                    "Estimated Budget Authority",
                    TextBlock(InlineFont(vertAlign="superscript"), "a"),
                ]
            )
            ws["D10"] = 103764
            workbook.save(raw / "51312-2024-06-snap.xlsx")

            rc = generate_schemas.generate_schemas(
                processed_dir=processed,
                schemas_dir=schemas,
                raw_dir=raw,
            )

            self.assertEqual(1, rc)

    def test_schema_file_name_matches_csv_basename(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"

            csv_names = ["snap_2024_06.csv", "medicaid_2026_02.csv", "child_nutrition_2019_05.csv"]
            for name in csv_names:
                _write_sample_csv(processed / name)

            generate_schemas.generate_schemas(processed_dir=processed, schemas_dir=schemas)

            for name in csv_names:
                expected = schemas / name.replace(".csv", ".md")
                self.assertTrue(expected.exists(), f"Expected schema file {expected} not found")

    def test_generate_schemas_returns_1_when_no_csvs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            rc = generate_schemas.generate_schemas(
                processed_dir=root / "empty",
                schemas_dir=root / "schemas",
            )
            self.assertEqual(1, rc)

    def test_readme_lists_all_datasets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            processed = root / "processed"
            schemas = root / "schemas"

            names = ["snap_2024_06.csv", "medicaid_2024_06.csv", "chip_2024_06.csv"]
            for name in names:
                _write_sample_csv(processed / name)

            generate_schemas.generate_schemas(processed_dir=processed, schemas_dir=schemas)

            readme = (schemas / "README.md").read_text(encoding="utf-8")
            for name in names:
                self.assertIn(name.replace(".csv", ""), readme)


if __name__ == "__main__":
    unittest.main()
