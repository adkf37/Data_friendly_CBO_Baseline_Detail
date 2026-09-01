import tempfile
import unittest
from csv import DictReader
from pathlib import Path

from openpyxl import Workbook

from etl import transform
from etl.config import get_output_path


def _release_path(output_dir: Path, workbook: str, output_dataset: str) -> Path:
    return get_output_path(
        output_dir,
        workbook=workbook,
        output_dataset=output_dataset,
    )


def _write_health_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Health"
    sheet["A1"] = "Health Program"
    sheet["B1"] = "2025"
    sheet["C1"] = "2026"
    sheet["A2"] = "Total Benefits"
    sheet["B2"] = 100
    sheet["C2"] = 110
    sheet["A3"] = "Administrative Costs"
    sheet["B3"] = 20
    sheet["C3"] = 22
    workbook.save(path)


def _write_income_security_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Income Security"
    sheet["A1"] = "Income Security Program"
    sheet["B1"] = "2025"
    sheet["C1"] = "2026"
    sheet["A2"] = "Total Benefits"
    sheet["B2"] = 100
    sheet["C2"] = 110
    sheet["A3"] = "Administrative Costs"
    sheet["B3"] = 20
    sheet["C3"] = 22
    workbook.save(path)


class TransformTests(unittest.TestCase):
    def test_child_nutrition_dataset_routes_to_remaining_programs_slice(self):
        plan = transform.SheetPlan(
            workbook="51293-2024-06-childnutrition.xlsx",
            sheet="CNP",
            include=True,
            output_dataset="childnutrition_2024_06",
            header_end_row=1,
            first_data_row=2,
            year_columns=[2, 3],
            unit="Millions of dollars",
        )

        self.assertFalse(transform._in_slice(plan, "health"))
        self.assertTrue(transform._in_slice(plan, "remaining-programs"))

    def test_run_transform_income_security_slice_excludes_health_datasets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_dir = root / "raw"
            output_dir = root / "processed"
            input_dir.mkdir(parents=True)

            health_workbook = "51293-2024-06-childnutrition.xlsx"
            income_workbook = "51312-2024-06-snap.xlsx"
            _write_health_workbook(input_dir / health_workbook)
            _write_income_security_workbook(input_dir / income_workbook)
            parse_plan_path = root / "workbook_parse_plan.yaml"
            parse_plan_path.write_text(
                """
workbooks:
  - workbook: 51293-2024-06-childnutrition.xlsx
    sheets:
      - sheet: Health
        include: true
        output_dataset: childnutrition_health_2024_06
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2, 3]
        unit: Millions of dollars
  - workbook: 51312-2024-06-snap.xlsx
    sheets:
      - sheet: Income Security
        include: true
        output_dataset: snap_2024_06
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2, 3]
        unit: Millions of dollars
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(
                parse_plan_path=parse_plan_path,
                input_dir=input_dir,
                output_dir=output_dir,
                slice_name="income-security",
            )

            self.assertEqual(0, rc)
            self.assertFalse(
                _release_path(output_dir, health_workbook, "childnutrition_health_2024_06").exists()
            )
            csv_path = _release_path(output_dir, income_workbook, "snap_2024_06")
            self.assertTrue(csv_path.exists())
            with csv_path.open(encoding="utf-8", newline="") as handle:
                rows = list(DictReader(handle))
            self.assertEqual(4, len(rows))
            self.assertEqual({"SNAP"}, {row["program"] for row in rows})
            self.assertEqual({"51312-2024-06-snap.xlsx"}, {row["source_file"] for row in rows})
            self.assertEqual({"Income Security"}, {row["source_sheet"] for row in rows})
            self.assertEqual({"Total Benefits", "Administrative Costs"}, {row["category"] for row in rows})
            self.assertEqual({"2025", "2026"}, {row["fiscal_year"] for row in rows})
            self.assertEqual("", (output_dir / "parse_errors.log").read_text(encoding="utf-8"))

    def test_run_transform_health_slice_writes_tidy_csv(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_dir = root / "raw"
            output_dir = root / "processed"
            input_dir.mkdir(parents=True)

            workbook_name = "51293-2024-06-childnutrition.xlsx"
            _write_health_workbook(input_dir / workbook_name)
            parse_plan_path = root / "workbook_parse_plan.yaml"
            parse_plan_path.write_text(
                """
workbooks:
  - workbook: 51293-2024-06-childnutrition.xlsx
    sheets:
      - sheet: Health
        include: true
        output_dataset: childnutrition_health_2024_06
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2, 3]
        unit: Millions of dollars
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(
                parse_plan_path=parse_plan_path,
                input_dir=input_dir,
                output_dir=output_dir,
                slice_name="health",
            )

            self.assertEqual(0, rc)
            csv_path = _release_path(
                output_dir, workbook_name, "childnutrition_health_2024_06"
            )
            self.assertTrue(csv_path.exists())
            with csv_path.open(encoding="utf-8", newline="") as handle:
                rows = list(DictReader(handle))
            self.assertEqual(transform.OUTPUT_COLUMNS, list(rows[0].keys()))
            self.assertEqual(4, len(rows))
            self.assertEqual(
                {
                    "program": "Child Nutrition",
                    "category": "Total Benefits",
                    "fiscal_year": "2025",
                    "value": "100.0",
                    "unit": "Millions of dollars",
                    "source_file": workbook_name,
                    "source_sheet": "Health",
                    "is_total": "true",
                    "program_id": "51293",
                    "category_path": "Total Benefits",
                    "period_type": "fiscal_year",
                    "period_start_year": "2025",
                    "period_end_year": "2025",
                    "period_label": "2025",
                    "source_row": "2",
                    "source_column": "2",
                },
                rows[0],
            )
            self.assertEqual("Administrative Costs", rows[-1]["category"])
            self.assertEqual("2026", rows[-1]["fiscal_year"])
            self.assertEqual("22.0", rows[-1]["value"])

            parse_errors = (output_dir / "parse_errors.log").read_text(encoding="utf-8")
            self.assertEqual("", parse_errors)

    def test_run_transform_infers_year_columns_when_missing_from_parse_plan(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_dir = root / "raw"
            output_dir = root / "processed"
            input_dir.mkdir(parents=True)

            workbook_name = "51293-2024-06-childnutrition.xlsx"
            _write_health_workbook(input_dir / workbook_name)
            parse_plan_path = root / "workbook_parse_plan.yaml"
            parse_plan_path.write_text(
                """
workbooks:
  - workbook: 51293-2024-06-childnutrition.xlsx
    sheets:
      - sheet: Health
        include: true
        output_dataset: childnutrition_health_2024_06
        header_rows: 1-1
        first_data_row: 2
        unit: Millions of dollars
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(
                parse_plan_path=parse_plan_path,
                input_dir=input_dir,
                output_dir=output_dir,
                slice_name="health",
            )

            self.assertEqual(0, rc)
            csv_path = _release_path(
                output_dir, workbook_name, "childnutrition_health_2024_06"
            )
            self.assertTrue(csv_path.exists())
            with csv_path.open(encoding="utf-8", newline="") as handle:
                rows = list(DictReader(handle))
            self.assertEqual(4, len(rows))
            self.assertEqual({"2025", "2026"}, {row["fiscal_year"] for row in rows})
            self.assertEqual("", (output_dir / "parse_errors.log").read_text(encoding="utf-8"))

    def test_run_transform_logs_parse_errors(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            parse_plan = root / "workbook_parse_plan.yaml"
            parse_plan.write_text(
                """
workbooks:
  - workbook: missing-health.xlsx
    sheets:
      - sheet: Missing
        include: true
        output_dataset: health_missing_2024
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2]
        unit: Millions
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(
                parse_plan_path=parse_plan,
                input_dir=root / "raw",
                output_dir=root / "processed",
                slice_name="health",
            )

            self.assertEqual(1, rc)
            error_text = (root / "processed" / "parse_errors.log").read_text(encoding="utf-8")
            self.assertIn("missing-health.xlsx", error_text)
            self.assertIn("workbook not found", error_text)

    def test_run_transform_remaining_programs_slice_excludes_health_and_income_security(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_dir = root / "raw"
            output_dir = root / "processed"
            input_dir.mkdir(parents=True)

            health_workbook = "51293-2024-06-childnutrition.xlsx"
            income_workbook = "51312-2024-06-snap.xlsx"
            other_workbook = "51315-2024-06-defense.xlsx"

            _write_health_workbook(input_dir / health_workbook)
            _write_income_security_workbook(input_dir / income_workbook)

            # Write a simple "other/remaining programs" workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Defense"
            ws["A1"] = "Defense Program"
            ws["B1"] = "2025"
            ws["C1"] = "2026"
            ws["A2"] = "Total Outlays"
            ws["B2"] = 800
            ws["C2"] = 850
            wb.save(input_dir / other_workbook)

            parse_plan = root / "workbook_parse_plan.yaml"
            parse_plan.write_text(
                """
workbooks:
  - workbook: 51293-2024-06-childnutrition.xlsx
    sheets:
      - sheet: Health
        include: true
        output_dataset: childnutrition_health_2024_06
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2, 3]
        unit: Millions of dollars
  - workbook: 51312-2024-06-snap.xlsx
    sheets:
      - sheet: Income Security
        include: true
        output_dataset: snap_2024_06
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2, 3]
        unit: Millions of dollars
  - workbook: 51315-2024-06-defense.xlsx
    sheets:
      - sheet: Defense
        include: true
        output_dataset: defense_2024_06
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2, 3]
        unit: Billions of dollars
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(
                parse_plan_path=parse_plan,
                input_dir=input_dir,
                output_dir=output_dir,
                slice_name="remaining-programs",
            )

            self.assertEqual(0, rc)
            self.assertFalse(
                _release_path(output_dir, health_workbook, "childnutrition_health_2024_06").exists()
            )
            self.assertFalse(_release_path(output_dir, income_workbook, "snap_2024_06").exists())
            csv_path = output_dir / "defense_2024_06.csv"
            self.assertTrue(csv_path.exists())
            with csv_path.open(encoding="utf-8", newline="") as handle:
                rows = list(DictReader(handle))
            self.assertEqual(2, len(rows))
            self.assertEqual({"Defense"}, {row["program"] for row in rows})
            self.assertEqual({"2025", "2026"}, {row["fiscal_year"] for row in rows})
            self.assertEqual("", (output_dir / "parse_errors.log").read_text(encoding="utf-8"))

    def test_run_transform_rejects_unknown_slice(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            with self.assertRaisesRegex(ValueError, "Unsupported slice: unknown"):
                transform.run_transform(
                    parse_plan_path=root / "workbook_parse_plan.yaml",
                    input_dir=root / "raw",
                    output_dir=root / "processed",
                    slice_name="unknown",
                )

    def test_run_transform_retains_historical_actual_year_columns(self):
        """Historical actuals are data and must not be silently discarded."""
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_dir = root / "raw"
            output_dir = root / "processed"
            input_dir.mkdir(parents=True)

            workbook_name = "51302-2019-05-Medicare.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Table 1"
            # Column 2 is a prior-year actual; columns 3-4 are projections.
            sheet["B1"] = 2018
            sheet["C1"] = 2019
            sheet["D1"] = 2020
            sheet["A2"] = "Total Outlays"
            sheet["B2"] = 700
            sheet["C2"] = 765
            sheet["D2"] = 814
            workbook.save(input_dir / workbook_name)

            parse_plan = root / "workbook_parse_plan.yaml"
            parse_plan.write_text(
                """
workbooks:
  - workbook: 51302-2019-05-Medicare.xlsx
    sheets:
      - sheet: Table 1
        include: true
        output_dataset: medicare_2019_05
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2, 3, 4]
        unit: Medicare Totals (Billions of dollars)
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(
                parse_plan_path=parse_plan,
                input_dir=input_dir,
                output_dir=output_dir,
                slice_name="health",
            )

            self.assertEqual(0, rc)
            csv_path = _release_path(output_dir, workbook_name, "medicare_2019_05")
            self.assertTrue(csv_path.exists())
            with csv_path.open(encoding="utf-8", newline="") as handle:
                rows = list(DictReader(handle))

            fiscal_years = [row["fiscal_year"] for row in rows]
            self.assertEqual({"2018", "2019", "2020"}, set(fiscal_years))
            self.assertEqual(3, len(rows))

    def test_run_transform_prefers_header_years_over_in_data_years(self):
        """Header-declared years are preferred over stale in-data years.

        Rows with the same category but *different* values represent distinct
        sub-programs and must both appear in the output.  All source rows are
        preserved; the transform does not deduplicate rows with coincidentally
        identical values.
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_dir = root / "raw"
            output_dir = root / "processed"
            input_dir.mkdir(parents=True)

            workbook_name = "51293-2024-06-childnutrition.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Health"
            sheet["A1"] = "Health Program"
            sheet["B1"] = "2024"
            sheet["C1"] = "2025"
            # Stale year value in the data area — must not become a fiscal year label.
            sheet["B2"] = 2096
            # Two distinct sub-program rows (different values → both kept).
            sheet["A3"] = "Paid"
            sheet["B3"] = 200
            sheet["C3"] = 210
            sheet["A4"] = "Free"
            sheet["B4"] = 50
            sheet["C4"] = 55
            # Row 5 has the same category and values as row 3; both rows are
            # written because each source row is preserved without deduplication.
            sheet["A5"] = "Paid"
            sheet["B5"] = 200
            sheet["C5"] = 210
            workbook.save(input_dir / workbook_name)

            parse_plan = root / "workbook_parse_plan.yaml"
            parse_plan.write_text(
                """
workbooks:
  - workbook: 51293-2024-06-childnutrition.xlsx
    sheets:
      - sheet: Health
        include: true
        output_dataset: childnutrition_health_2024_06
        header_rows: 1-4
        first_data_row: 3
        year_columns: [2, 3]
        unit: Millions
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(
                parse_plan_path=parse_plan,
                input_dir=input_dir,
                output_dir=output_dir,
                slice_name="health",
            )

            self.assertEqual(0, rc)
            csv_path = _release_path(
                output_dir, workbook_name, "childnutrition_health_2024_06"
            )
            with csv_path.open(encoding="utf-8", newline="") as handle:
                rows = list(DictReader(handle))

            # 3 source rows × 2 fiscal years = 6 rows; rows 3 and 5 both appear
            # since every source row is preserved regardless of value equality.
            self.assertEqual(6, len(rows))
            fiscal_years = [row["fiscal_year"] for row in rows]
            self.assertEqual(["2024", "2025", "2024", "2025", "2024", "2025"], fiscal_years)
            values = [row["value"] for row in rows]
            self.assertEqual(["200.0", "210.0", "50.0", "55.0", "200.0", "210.0"], values)
            self.assertNotIn("2096", fiscal_years)

    def test_extract_years_scans_beyond_header_end_row(self):
        """Years in a row below the declared header_end_row are still found."""
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Data"
        # Row 1 (header): category label only — no year values
        ws["A1"] = "Category"
        # Row 4: year values as integers (below header_rows: 1-1)
        ws["B4"] = 2020
        ws["C4"] = 2021
        # Row 5: data
        ws["A5"] = "Budget Authority"
        ws["B5"] = 100
        ws["C5"] = 110

        plan = transform.SheetPlan(
            workbook="test.xlsx",
            sheet="Data",
            include=True,
            output_dataset="test_dataset",
            header_end_row=1,  # declared header covers only row 1
            first_data_row=5,
            year_columns=[2, 3],
            unit="Millions of dollars",
        )

        years = transform._extract_years(ws, plan)
        self.assertEqual({2: 2020, 3: 2021}, years, "Years below declared header_end_row must be found")

    def test_extract_years_skips_datetime_cells(self):
        """Datetime objects (publication dates) in year columns must not be treated as fiscal years."""
        import datetime as dt

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Data"
        # Row 1: a publication date datetime in column 2 — must not become a fiscal year label
        ws.cell(row=1, column=2).value = dt.datetime(2020, 5, 2)
        # Row 2: actual integer year value in column 2
        ws.cell(row=2, column=2).value = 2025
        ws.cell(row=3, column=1).value = "Program Detail"
        ws.cell(row=3, column=2).value = 500

        plan = transform.SheetPlan(
            workbook="test.xlsx",
            sheet="Data",
            include=True,
            output_dataset="test_dataset",
            header_end_row=1,
            first_data_row=3,
            year_columns=[2],
            unit="Millions",
        )

        years = transform._extract_years(ws, plan)
        # The datetime in row 1 must be skipped; row 2 (integer 2025) is not in
        # the declared header, but MAX_YEAR_SCAN_ROWS extends the search.
        self.assertEqual({2: 2025}, years, "datetime cell must be ignored; integer year row must be found")

    def test_transform_marks_stacked_multi_year_columns_as_cumulative(self):
        workbook = Workbook()
        ws = workbook.active
        ws["E1"] = "2025-"
        ws["B2"] = 2024
        ws["C2"] = 2025
        ws["D2"] = 2026
        ws["E2"] = 2029
        ws["A3"] = "Estimated Outlays"
        ws["B3"] = 10
        ws["C3"] = 11
        ws["D3"] = 12
        ws["E3"] = 55
        plan = transform.SheetPlan(
            workbook="51293-2024-06-childnutrition.xlsx",
            sheet="Data",
            include=True,
            output_dataset="test",
            header_end_row=2,
            first_data_row=3,
            year_columns=[2, 3, 4, 5],
            unit="Millions of dollars",
        )

        rows, warning = transform._records_for_sheet(ws, plan)

        cumulative = next(row for row in rows if row["source_column"] == 5)
        self.assertIsNone(warning)
        self.assertEqual("cumulative_fiscal_years", cumulative["period_type"])
        self.assertEqual("", cumulative["fiscal_year"])
        self.assertEqual(2025, cumulative["period_start_year"])
        self.assertEqual(2029, cumulative["period_end_year"])
        self.assertEqual("2025-2029", cumulative["period_label"])

    def test_transform_honors_calendar_year_period_hint(self):
        workbook = Workbook()
        ws = workbook.active
        ws["D1"] = 2025
        ws["E1"] = 2026
        ws["A2"] = "Covered Population"
        ws["D2"] = 100
        ws["E2"] = 110
        plan = transform.SheetPlan(
            workbook="51298-2026-02-healthinsurance.xlsx",
            sheet="Coverage",
            include=True,
            output_dataset="healthinsurance_2026_02",
            header_end_row=1,
            first_data_row=2,
            year_columns=[4, 5],
            unit="Millions of people",
            period_type="calendar_year",
        )

        rows, warning = transform._records_for_sheet(ws, plan)

        self.assertIsNone(warning)
        self.assertEqual({"calendar_year"}, {row["period_type"] for row in rows})
        self.assertEqual({""}, {row["fiscal_year"] for row in rows})

    def test_transform_labels_heading_only_rows_and_excludes_note_block_numbers(self):
        workbook = Workbook()
        ws = workbook.active
        ws["A1"] = "Coverage Share"
        ws["D2"] = 2025
        ws["E2"] = 2026
        ws["D3"] = 30
        ws["E3"] = 31
        ws["A4"] = "Note:"
        ws["D5"] = 4
        ws["E5"] = 5
        ws["A6"] = "Illustrative thresholds"
        ws["D6"] = 6
        ws["E6"] = 7
        plan = transform.SheetPlan(
            workbook="test.xlsx",
            sheet="Data",
            include=True,
            output_dataset="test",
            header_end_row=2,
            first_data_row=3,
            year_columns=[4, 5],
            unit="Percent",
        )

        rows, warning = transform._records_for_sheet(ws, plan)

        self.assertIsNone(warning)
        self.assertEqual(2, len(rows))
        self.assertEqual({3}, {row["source_row"] for row in rows})
        self.assertEqual({"Coverage Share"}, {row["category"] for row in rows})

    def test_normalize_unit_removes_concatenated_superscript_marker(self):
        self.assertEqual(
            "Number of beneficiaries",
            transform._normalize_unit_string("Number of Beneficiariese"),
        )

    def test_usda_hierarchy_splits_two_three_and_four_level_paths(self):
        cases = [
            (
                "TABLE / Leaf",
                "Leaf",
                {"table_title": "TABLE", "section": "", "subsection": ""},
            ),
            (
                "TABLE / Supply / Production",
                "Production",
                {"table_title": "TABLE", "section": "Supply", "subsection": ""},
            ),
            (
                "TABLE / Program / Component / Outlays",
                "Outlays",
                {
                    "table_title": "TABLE",
                    "section": "Program",
                    "subsection": "Component",
                },
            ),
        ]

        for category_path, category, expected in cases:
            with self.subTest(category_path=category_path):
                self.assertEqual(expected, transform._usda_hierarchy(category_path, category))

    def test_usda_record_and_csv_include_hierarchy_columns(self):
        plan = transform.SheetPlan(
            workbook="51317-2026-02-usda.xlsx",
            sheet="USDA Baseline_02-2026",
            include=True,
            output_dataset="usda_2026_02",
            header_end_row=1,
            first_data_row=2,
            year_columns=[4],
            unit="Millions of dollars",
        )
        record = transform._record(
            plan=plan,
            category="Production",
            category_path="CORN SUPPLY AND USE / Supply / Production",
            period_type="fiscal_year",
            start_year=2026,
            end_year=2026,
            period_label="2026",
            value=100.0,
            unit="Millions",
            row=10,
            column=4,
        )

        self.assertEqual("CORN SUPPLY AND USE", record["table_title"])
        self.assertEqual("Supply", record["section"])
        self.assertEqual("", record["subsection"])

        with tempfile.TemporaryDirectory() as tmpdir:
            csv_path = Path(tmpdir) / "usda_2026_02.csv"
            transform._write_dataset(csv_path, [record])
            with csv_path.open(encoding="utf-8", newline="") as handle:
                reader = DictReader(handle)
                rows = list(reader)
                self.assertEqual(transform.USDA_OUTPUT_COLUMNS, list(reader.fieldnames or []))
            self.assertEqual("CORN SUPPLY AND USE", rows[0]["table_title"])
            self.assertEqual("Supply", rows[0]["section"])

    def test_transform_recognizes_award_year_ranges(self):
        workbook = Workbook()
        ws = workbook.active
        ws["A1"] = "Percent, by award year"
        ws["D2"] = "2024-25"
        ws["E2"] = "2025-26"
        ws["A3"] = "Undergraduate Loans"
        ws["D3"] = 6.5
        ws["E3"] = 6.2
        plan = transform.SheetPlan(
            workbook="51310-2024-06-studentloan.xlsx",
            sheet="Data",
            include=True,
            output_dataset="test",
            header_end_row=2,
            first_data_row=3,
            year_columns=[],
            unit="Percent",
        )

        rows, warning = transform._records_for_sheet(ws, plan)

        self.assertIsNone(warning)
        self.assertEqual({"award_year"}, {row["period_type"] for row in rows})
        self.assertEqual({(2024, 2025), (2025, 2026)}, {
            (row["period_start_year"], row["period_end_year"]) for row in rows
        })
        self.assertEqual({""}, {row["fiscal_year"] for row in rows})

    def test_transform_uses_row_metric_unit_and_preserves_hierarchy(self):
        workbook = Workbook()
        ws = workbook.active
        ws["D1"] = 2025
        ws["E1"] = 2026
        ws["A2"] = "Direct Loans"
        ws["A3"] = "Program Account"
        ws["C3"] = "Benefit per Borrower (Dollars)"
        ws["D3"] = 100
        ws["E3"] = 110
        plan = transform.SheetPlan(
            workbook="51310-2024-06-studentloan.xlsx",
            sheet="Data",
            include=True,
            output_dataset="test",
            header_end_row=1,
            first_data_row=3,
            year_columns=[4, 5],
            unit="Billions of dollars",
        )

        rows, _ = transform._records_for_sheet(ws, plan)

        self.assertEqual({"Dollars"}, {row["unit"] for row in rows})
        self.assertEqual(
            {"Direct Loans / Program Account / Benefit per Borrower (Dollars)"},
            {row["category_path"] for row in rows},
        )

    def test_transform_includes_formerly_verification_exempt_sheet(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            raw_dir = root / "raw"
            output_dir = root / "processed"
            raw_dir.mkdir()
            workbook_name = "51298-2020-03-healthinsurance.xlsx"
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Table 4"
            ws["B1"] = 2025
            ws["C1"] = 2026
            ws["A2"] = "Enrollment"
            ws["B2"] = 10
            ws["C2"] = 11
            workbook.save(raw_dir / workbook_name)
            parse_plan = root / "plan.yaml"
            parse_plan.write_text(
                """
workbooks:
  - workbook: 51298-2020-03-healthinsurance.xlsx
    sheets:
      - sheet: Table 4
        include: true
        output_dataset: healthinsurance_2020_03
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2, 3]
        unit: Millions of people
        verification_exempt: true
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(parse_plan, raw_dir, output_dir, "health")

            self.assertEqual(0, rc)
            self.assertTrue(
                _release_path(output_dir, workbook_name, "healthinsurance_2020_03").exists()
            )

    def test_find_sheet_exact_match(self):
        self.assertEqual("MySheet", transform._find_sheet(["MySheet", "Other"], "MySheet"))

    def test_find_sheet_trailing_space_match(self):
        """Sheet names with trailing spaces in the workbook are matched against the parse plan entry."""
        self.assertEqual("MySheet ", transform._find_sheet(["MySheet ", "Other"], "MySheet"))

    def test_find_sheet_returns_none_when_not_found(self):
        self.assertIsNone(transform._find_sheet(["Other", "Another"], "Missing"))

    def test_run_transform_handles_trailing_space_sheet_name(self):
        """Transform succeeds when the workbook sheet has a trailing space not in the parse plan."""
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            input_dir = root / "raw"
            output_dir = root / "processed"
            input_dir.mkdir(parents=True)

            workbook_name = "51297-2020-03-mortgages.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Mortgage Programs "  # trailing space
            ws["A1"] = "Category"
            ws["B1"] = "2025"
            ws["A2"] = "Total Outlays"
            ws["B2"] = 500
            wb.save(input_dir / workbook_name)

            parse_plan = root / "workbook_parse_plan.yaml"
            parse_plan.write_text(
                """
workbooks:
  - workbook: 51297-2020-03-mortgages.xlsx
    sheets:
      - sheet: Mortgage Programs
        include: true
        output_dataset: mortgages_2020_03
        header_rows: 1-1
        first_data_row: 2
        year_columns: [2]
        unit: Millions of dollars
                """.strip()
                + "\n",
                encoding="utf-8",
            )

            rc = transform.run_transform(
                parse_plan_path=parse_plan,
                input_dir=input_dir,
                output_dir=output_dir,
                slice_name="remaining-programs",
            )

            self.assertEqual(0, rc)
            csv_path = _release_path(output_dir, workbook_name, "mortgages_2020_03")
            self.assertTrue(csv_path.exists(), "CSV should be written even when sheet name has trailing space")


if __name__ == "__main__":
    unittest.main()
