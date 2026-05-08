import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src import inspect


def _write_fixture_workbook(path: Path) -> None:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Health Spending"
    data_sheet["A1"] = "Program baseline (billions of dollars)"
    data_sheet.merge_cells("A1:C1")
    data_sheet["A2"] = "Category"
    data_sheet["B2"] = "2025"
    data_sheet["C2"] = "2026"
    data_sheet["A3"] = "Medicare"
    data_sheet["B3"] = 100
    data_sheet["C3"] = 110
    data_sheet["A7"] = "Medicaid"
    data_sheet["B7"] = 120
    data_sheet["C7"] = 130
    data_sheet["A8"] = "CHIP"
    data_sheet["B8"] = 20
    data_sheet["C8"] = 22

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet["A1"] = "Notes and source text"
    notes_sheet["A2"] = "Source: CBO"

    workbook.save(path)


class InspectTests(unittest.TestCase):
    def test_profile_sheet_detects_expected_fields(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "health.xlsx"
            _write_fixture_workbook(workbook_path)

            profile = inspect.profile_sheet(workbook_path, "Health Spending")

            self.assertEqual("Health Spending", profile["sheet_name"])
            self.assertEqual(3, profile["column_count"])
            self.assertTrue(profile["has_merged_cells"])
            self.assertEqual([2, 3], profile["fiscal_year_columns"])
            self.assertEqual("data", profile["classification"])
            self.assertTrue(profile["multiple_tables_flagged"])
            self.assertEqual(3, profile["likely_first_data_row"])

    def test_run_inspection_writes_report(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = Path(tmpdir) / "raw"
            input_dir.mkdir(parents=True)
            workbook_path = input_dir / "health.xlsx"
            _write_fixture_workbook(workbook_path)

            report_path = Path(tmpdir) / "docs" / "inspection_report.md"
            rc = inspect.run_inspection(input_dir=input_dir, report_path=report_path)

            self.assertEqual(0, rc)
            self.assertTrue(report_path.exists())
            text = report_path.read_text(encoding="utf-8")
            self.assertIn("## health.xlsx", text)
            self.assertIn("### Health Spending", text)
            self.assertIn("### Notes", text)
            self.assertIn("Multiple logical tables flagged: yes", text)
            self.assertIn("Classification: `notes`", text)


if __name__ == "__main__":
    unittest.main()
