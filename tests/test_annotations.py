import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from etl import annotations as source_annotations


def _rich_label(base: str, markers: str) -> CellRichText:
    return CellRichText(
        [
            base,
            TextBlock(InlineFont(vertAlign="superscript"), markers),
        ]
    )


class SourceAnnotationTests(unittest.TestCase):
    def test_extract_superscript_label_uses_actual_rich_text(self):
        workbook = Workbook()
        ws = workbook.active
        ws["A1"] = _rich_label("Budget Authority", "a")

        result = source_annotations.extract_superscript_label(ws["A1"])

        self.assertIsNotNone(result)
        self.assertEqual("Budget Authority", result.base_label)
        self.assertEqual("Budget Authoritya", result.raw_label)
        self.assertEqual(("a",), result.markers)

    def test_extract_superscript_label_supports_comma_separated_markers(self):
        workbook = Workbook()
        ws = workbook.active
        ws["A1"] = _rich_label("Medicare", "d,h")

        result = source_annotations.extract_superscript_label(ws["A1"])

        self.assertIsNotNone(result)
        self.assertEqual(("d", "h"), result.markers)
        self.assertEqual("Medicare", result.base_label)

    def test_extract_worksheet_annotations_keeps_wrapped_note_text(self):
        workbook = Workbook()
        ws = workbook.active
        ws["A2"] = _rich_label("Budget Authority", "a")
        ws["B2"] = 10
        ws["A5"] = "a."
        ws["B5"] = "The first sentence."
        ws["A6"] = "The wrapped second sentence."

        annotations = source_annotations.extract_worksheet_annotations(ws)

        self.assertEqual(1, len(annotations))
        self.assertEqual(5, annotations[0].note_row)
        self.assertEqual(
            "The first sentence. The wrapped second sentence.",
            annotations[0].variable_note,
        )

    def test_repeated_marker_uses_nearest_following_definition(self):
        workbook = Workbook()
        ws = workbook.active
        ws["A2"] = _rich_label("First table item", "a")
        ws["B2"] = 10
        ws["A4"] = "a. First table note."
        ws["A8"] = _rich_label("Second table item", "a")
        ws["B8"] = 20
        ws["A10"] = "a. Second table note."

        annotations = source_annotations.extract_worksheet_annotations(ws)

        self.assertEqual(2, len(annotations))
        self.assertEqual("First table note.", annotations[0].variable_note)
        self.assertEqual("Second table note.", annotations[1].variable_note)

    def test_missing_definition_remains_unresolved_for_catalog_audit(self):
        workbook = Workbook()
        ws = workbook.active
        ws["A2"] = _rich_label("Inflated Budget Authority", "d")
        ws["B2"] = 10

        annotations = source_annotations.extract_worksheet_annotations(ws)

        self.assertEqual(1, len(annotations))
        self.assertIsNone(annotations[0].note_row)
        self.assertEqual("", annotations[0].variable_note)

    def test_catalog_documents_known_missing_pell_definition(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            raw_dir = Path(tmpdir)
            workbook = Workbook()
            ws = workbook.active
            ws.title = "T4_Pell_02-2026"
            ws["A30"] = _rich_label("Cumulative Shortfall Inflated", "d")
            ws["D31"] = 10
            workbook.save(raw_dir / "51304-2026-02-pellgrant.xlsx")

            catalog = source_annotations.load_annotation_catalog(
                raw_dir,
                {"51304-2026-02-pellgrant.xlsx": {"T4_Pell_02-2026"}},
            )

            loaded = catalog.by_source[
                ("51304-2026-02-pellgrant.xlsx", "T4_Pell_02-2026")
            ]
            self.assertEqual(1, catalog.resolved_marker_references)
            self.assertIsNone(loaded[0].note_row)
            self.assertEqual(
                "The source workbook includes superscript marker 'd' but does not "
                "provide a corresponding footnote definition on this worksheet.",
                loaded[0].variable_note,
            )

    def test_parent_heading_note_applies_to_child_category_path(self):
        workbook = Workbook()
        ws = workbook.active
        ws["A2"] = _rich_label("Enrollment", "b")
        ws["A3"] = "Children"
        ws["B3"] = 7
        ws["A5"] = "b. Average monthly enrollment."
        annotations = source_annotations.extract_worksheet_annotations(ws)
        catalog = source_annotations.AnnotationCatalog(
            by_source={("sample.xlsx", "Data"): annotations}
        )
        contexts = [
            source_annotations.ObservationContext(
                source_file="sample.xlsx",
                source_sheet="Data",
                source_row=3,
                source_column=2,
                category_path="Enrollmentb / Children",
            )
        ]

        notes = source_annotations.match_variable_notes(contexts, catalog)

        self.assertEqual(1, len(notes))
        self.assertEqual("b", notes[0].marker)
        self.assertEqual("Average monthly enrollment.", notes[0].variable_note)

    def test_spanning_header_note_applies_by_source_region(self):
        annotation = source_annotations.ResolvedAnnotation(
            label_row=2,
            label_column=2,
            raw_label="Insurance Coveragea",
            base_label="Insurance Coverage",
            marker="a",
            note_row=10,
            variable_note="Coverage categories can overlap.",
        )
        catalog = source_annotations.AnnotationCatalog(
            by_source={("sample.xlsx", "Data"): (annotation,)}
        )
        contexts = [
            source_annotations.ObservationContext(
                source_file="sample.xlsx",
                source_sheet="Data",
                source_row=4,
                source_column=3,
                category_path="Employment-Based Coverage",
            )
        ]

        notes = source_annotations.match_variable_notes(contexts, catalog)

        self.assertEqual(1, len(notes))
        self.assertEqual("Coverage categories can overlap.", notes[0].variable_note)

    def test_spanning_header_keeps_every_marker_from_same_cell(self):
        annotations = tuple(
            source_annotations.ResolvedAnnotation(
                label_row=2,
                label_column=2,
                raw_label="Coveragea,b",
                base_label="Coverage",
                marker=marker,
                note_row=10 + index,
                variable_note=f"Note {marker}.",
            )
            for index, marker in enumerate(("a", "b"))
        )
        catalog = source_annotations.AnnotationCatalog(
            by_source={("sample.xlsx", "Data"): annotations}
        )
        contexts = [
            source_annotations.ObservationContext(
                source_file="sample.xlsx",
                source_sheet="Data",
                source_row=4,
                source_column=3,
                category_path="Coverage detail",
            )
        ]

        notes = source_annotations.match_variable_notes(contexts, catalog)

        self.assertEqual({"a", "b"}, {note.marker for note in notes})

    def test_base_label_does_not_contaminate_later_unannotated_reuse(self):
        annotation = source_annotations.ResolvedAnnotation(
            label_row=2,
            label_column=1,
            raw_label="Annual Subsidy Costsc",
            base_label="Annual Subsidy Costs",
            marker="c",
            note_row=20,
            variable_note="Applies only to the annotated program.",
        )
        catalog = source_annotations.AnnotationCatalog(
            by_source={("sample.xlsx", "Data"): (annotation,)}
        )
        contexts = [
            source_annotations.ObservationContext(
                source_file="sample.xlsx",
                source_sheet="Data",
                source_row=2,
                source_column=2,
                category_path="First Program / Annual Subsidy Costsc",
            ),
            source_annotations.ObservationContext(
                source_file="sample.xlsx",
                source_sheet="Data",
                source_row=8,
                source_column=2,
                category_path="Second Program / Annual Subsidy Costs",
            ),
        ]

        notes = source_annotations.match_variable_notes(contexts, catalog)

        self.assertEqual(1, len(notes))
        self.assertEqual("First Program / Annual Subsidy Costsc", notes[0].category_path)

    def test_distant_unmatched_label_is_retained_as_source_only(self):
        annotation = source_annotations.ResolvedAnnotation(
            label_row=2,
            label_column=1,
            raw_label="Unparsed itema",
            base_label="Unparsed item",
            marker="a",
            note_row=20,
            variable_note="Only the unparsed item is affected.",
        )
        catalog = source_annotations.AnnotationCatalog(
            by_source={("sample.xlsx", "Data"): (annotation,)}
        )
        contexts = [
            source_annotations.ObservationContext(
                source_file="sample.xlsx",
                source_sheet="Data",
                source_row=8,
                source_column=2,
                category_path="Different table item",
            )
        ]

        notes = source_annotations.match_variable_notes(contexts, catalog)

        self.assertEqual(1, len(notes))
        self.assertEqual("", notes[0].category_path)
        self.assertEqual("Unparsed itema", notes[0].source_label)

    def test_catalog_loads_rich_text_from_saved_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            raw_dir = Path(tmpdir)
            workbook = Workbook()
            ws = workbook.active
            ws.title = "Data"
            ws["A2"] = _rich_label("Outlays", "a,b")
            ws["B2"] = 5
            ws["A4"] = "a. First note."
            ws["A5"] = "b. Second note."
            workbook.save(raw_dir / "sample.xlsx")

            catalog = source_annotations.load_annotation_catalog(
                raw_dir,
                {"sample.xlsx": {"Data"}},
            )

            loaded = catalog.by_source[("sample.xlsx", "Data")]
            self.assertEqual(2, len(loaded))
            self.assertEqual(2, catalog.resolved_marker_references)
            self.assertEqual({"First note.", "Second note."}, {item.variable_note for item in loaded})


if __name__ == "__main__":
    unittest.main()
